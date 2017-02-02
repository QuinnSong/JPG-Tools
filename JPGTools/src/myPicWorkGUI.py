# -*- coding: cp936 -*- 

###########################################################################
## Python code generated with wxFormBuilder (version Jun  5 2014)
## http://www.wxformbuilder.org/
###########################################################################

import wx, os
from PIL import Image, ImageEnhance, ImageFilter, ImageDraw, ImageFont
from wx.lib.pubsub import pub
from round import round_image
from background import background, FILL_CHOICES
from shadow import drop_shadow, invert
from border_v2 import border
from molten import molten
from lighting import lighting
from reflection import reflect
from mirror import tile
from sketch import sketch
from ice import ice
from papercut import paper_cut
from tan import tan
from swirl import swirl
from spherize import spherize
#from wave import wave
from StringIO import StringIO
from de_yellow import deYellow
from np_puzzle import ImgSalon
from light import edge_blur
from corner_fold import EdgeFold
from tear import Tear
from Fade import Fade
from mix import MixColor
from color_replace import Replace
from np_pil import newPil
from auto_brightness import equalize, brightness
from shear import Shear
from img_mosaic import mosaic
from blendOrMask import AddMask, blend
from split import split
from slice_gone import slice_gone
import iconfile
import numpy as np
import _winreg
from openpyxl import load_workbook, workbook, worksheet
from openpyxl.utils.exceptions import InvalidFileException
import math

COLOR_FORMAT_TIP = u"�ֶ�����ʱ���밴�ո�ʽ��\r\n(192, 192, 192, 255) ���� #FF3CD2\r\n���⣬��Ҳ����������������ı���Ȼ���ٵ��ͼƬ���ȡɫ"
BACKGROUND_TIP = u'�򿪱�������\r\nΪ͸��ͼƬ��ӱ���'
PUZZLE_TIP = u"��ƴͼ����\r\n����С��ͬ��ͼƬƴ��һ��"
SPLIT_TIP = u"��ͼƬ�ָ�����\r\n����ǰͼƬ�ָ�ɴ�С��ͬ��ͼƬ��\r\nע�⣺�����Ӹ������������һ��ִ��"
FOCUS_TIP = u'�򿪾۹�����\r\n��ʾ����Ȧ����Ȧ�뾶�������ͼƬ��Բ����'
TEAR_TIP = u"�򿪾������\r\n��ָ���������ɾ��Ч����"
REPLACE_TIP = u"����ɫ�滻����\r\n��ͼƬ������������һ����ɫ����һ�����滻"
SPLIT_LINE_COLOR_TIP = u"Ԥ��ʱ�ķָ�����ɫ"
SLICE_TIP = u"�����Ƚ���Ҫȥ�����м䲿�֣������ѡ�С����Խ������»�����ƴ��"

wildcard = "�����ļ� (*.*) | *.*|" \
	"JPG ͼƬ (*.jpg) | *.jpg|" \
	"BMP ͼƬ (*.bmp) | *.bmp|" \
	"PNG ͼƬ (*.png) | *.png"

###########################################################################
## Class MyJPGWorkDlg
###########################################################################

class MyJPGWorkDlg ( wx.Dialog ):
	
	def __init__( self, parent ):
		wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"ͼƬ��Ч", pos = wx.DefaultPosition,
				    size = wx.Size( 570, 820 ), style = wx.DEFAULT_DIALOG_STYLE |wx.MINIMIZE_BOX | wx.TE_PROCESS_ENTER )				
		self.parent = parent
		self.list_view_dlg = None
		self.SetIcon(self.parent.midi)	
		self.ctrlOnFocus = None
		self.settingCounts = 0
		self.out_dir = None
		#self.lastBtnPressed = None
		self.cur_im = None # the current Image object
		self.cur_img = None # the image current working on
		#self.txtDlg = None # watermark instance
		self.taskMapper = {u"Բ��Ч��": self.MakeRound, u"����Ч��": self.MakeBackground, \
				   u"�߿�Ч��": self.MakeBorder, u"��ӰЧ��": self.MakeShadow, \
				   u"��ӰЧ��": self.MakeReflection, u"�Ҷ�Ч��": self.MakeGrayscale, \
				   
				   u"�ڰ�Ч��": self.MakeBlkWht, u"ͼ����ǿ": self.MakeEnhance, \
				   u"�˾�Ч��": self.MakeFilter, u"��תЧ��": self.MakeRotate, \
				   u"����ת": self.MakeMirror, u"��ƬЧ��": self.MakeInvert, \
				   
				   u"����Ч��": self.MakeMolten, u"�ƹ�Ч��":  self.MakeLighting, \
				   u"����Ч��": self.MakeSketch, u"����Ч��": self.MakeIce, \
				   u"��ֽЧ��": self.MakePapercut, u"�۹�Ч��": self.MakeFocus,\
				   
				   u"����Ƭ": self.MakeDeYellow, u"����Ƭ": self.MakeTan, \
				   u"����Ч��": self.MakeSwirl, u"������": self.MakeSpherize, \
				   u"ƴͼЧ��": self.MakePuzzle, u"�۽�Ч��": self.MakeCornerFold, \
				   u"��ݱ�Ե": self.MakeTear, u"����Ч��": self.MakeFade, \
				   u"��ɫЧ��": self.MakeMix, u"��ɫ�滻": self.MakeReplace, \
				   u"�Զ�����": self.MakeAutoBrightness, u"����Ч��": self.MakeRoll, \
				   u"��бЧ��": self.MakeShear, u"������": self.MakeMosaic, \
				   u"�ɰ�Ч��": self.AddMask, u"���Ч��": self.MakeBlend, \
				   u"ˮӡЧ��": self.MakeWaterMark, u"ͼƬ�ָ�": self.MakeSplit, \
				   u"�м�ȥ��": self.SliceGone }
		
		bSizer1 = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_panel1 = wx.Panel( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL )
		bSizerMain = wx.BoxSizer( wx.VERTICAL )
		
		bSizerTask = wx.BoxSizer( wx.HORIZONTAL )
		
		bSizerTaskSub = wx.BoxSizer( wx.HORIZONTAL )            
		sbSizerTaskList = wx.StaticBoxSizer( wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�����б�����" ), wx.VERTICAL )         
		self.m_checkList1Choices = []
		self.m_checkList1 = wx.ListBox( self.m_panel1, wx.ID_ANY, wx.DefaultPosition, wx.Size( 440,200 ), self.m_checkList1Choices, wx.LB_NEEDED_SB|wx.LB_SINGLE )
		sbSizerTaskList.Add( self.m_checkList1, 0, wx.ALL|wx.EXPAND, 5 )
		self.m_checkList1.Bind(wx.EVT_LISTBOX, self.KeyDown)
		self.m_checkList1.SetToolTipString(u'���"+" ��ť���Ч�����б�')            
		bSizerTask.Add( sbSizerTaskList, 1, wx.EXPAND, 5 )
		
		bSizerUpDown = wx.BoxSizer( wx.VERTICAL )
		self.m_buttonDel = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.remove_ico, wx.DefaultPosition, wx.Size(24, 24), wx.BU_EXACTFIT  )
		bSizerUpDown.Add( self.m_buttonDel, 0, wx.ALL | wx.CENTER, 5 )
		self.m_buttonDel.SetToolTipString(u'�Ƴ���ѡ����')		
		
		self.m_buttonDelAll = wx.Button( self.m_panel1, wx.ID_ANY, "", wx.DefaultPosition, wx.Size(24, 24), 0 )
		bSizerUpDown.Add( self.m_buttonDelAll, 0, wx.ALL |  wx.CENTER, 5 )
		self.m_buttonDelAll.SetToolTipString(u'��������б�')         
		bSizerUpDown.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )		
		
		self.m_btnUP = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.up_ico, wx.DefaultPosition, wx.Size(24, 24), wx.BU_EXACTFIT  )
		bSizerUpDown.Add( self.m_btnUP, 0, wx.ALL | wx.CENTER, 5 )
		self.m_btnUP.SetToolTipString(u'����')

		self.m_btnDown = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.down_ico, wx.DefaultPosition, wx.Size(24, 24), wx.BU_EXACTFIT  )
		bSizerUpDown.Add( self.m_btnDown, 0, wx.ALL |  wx.CENTER, 5 )
		self.m_btnDown.SetToolTipString(u'����')		
		
		bSizerTask.Add( bSizerUpDown, 1, wx.EXPAND, 5 )
		bSizerMain.Add( bSizerTask, 1, wx.EXPAND, 5 )
		
		# Create a panel for all available effects ---------------              
		sbSizerTaskPanel = wx.StaticBoxSizer( wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��ѡЧ��"), wx.VERTICAL )
		sbSizerTaskPanel_1 = wx.BoxSizer(wx.HORIZONTAL)
		sbSizerTaskPanel_2 = wx.BoxSizer(wx.HORIZONTAL)
		sbSizerTaskPanel_3 = wx.BoxSizer(wx.HORIZONTAL)
		sbSizerTaskPanel_4 = wx.BoxSizer(wx.HORIZONTAL)
		sbSizerTaskPanel_5 = wx.BoxSizer(wx.HORIZONTAL)
		sbSizerTaskPanel_6 = wx.BoxSizer(wx.HORIZONTAL)
		sbSizerTaskPanel_7 = wx.BoxSizer(wx.HORIZONTAL)
		
		self.m_buttonRound = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"Բ��Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_1.Add( self.m_buttonRound, 0, wx.ALL, 5 )
		#self.m_buttonRound.SetValue(True)
		self.m_buttonRound.SetToolTipString(u'��Բ������')
		
		self.m_buttonBackground = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_1.Add( self.m_buttonBackground, 0, wx.ALL, 5 )
		self.m_buttonBackground.SetToolTipString(BACKGROUND_TIP)
		
		self.m_buttonBorder = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�߿�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_1.Add( self.m_buttonBorder, 0, wx.ALL, 5 )
		self.m_buttonBorder.SetToolTipString(u'�򿪱߿�����')
		## 
		self.m_buttonShadow = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ӰЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_1.Add( self.m_buttonShadow, 0, wx.ALL, 5 )
		self.m_buttonShadow.SetToolTipString(u'����Ӱ����')
		
		self.m_buttonReflection = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ӰЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_1.Add( self.m_buttonReflection, 0, wx.ALL, 5 )
		self.m_buttonReflection.SetToolTipString(u'�򿪵�Ӱ����')
		
		self.m_buttonGrayscale = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�Ҷ�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_1.Add( self.m_buttonGrayscale, 0, wx.ALL, 5 )		
		
		self.m_buttonBlkWht = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�ڰ�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_2.Add( self.m_buttonBlkWht, 0, wx.ALL, 5 )     
		# ��ǿ
		self.m_buttonEnhance = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"ͼ����ǿ", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_2.Add( self.m_buttonEnhance, 0, wx.ALL, 5 )
		self.m_buttonEnhance.SetToolTipString(u'��ͼ����ǿ����')        
		
		self.m_buttonFilter = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�˾�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_2.Add( self.m_buttonFilter, 0, wx.ALL, 5 )
		self.m_buttonFilter.SetToolTipString(u'���˾�����')
		
		self.m_buttonRotate = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��תЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_2.Add( self.m_buttonRotate, 0, wx.ALL, 5 )
		self.m_buttonRotate.SetToolTipString(u'����ת����')
		
		self.m_buttonMirror = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����ת", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_2.Add( self.m_buttonMirror, 0, wx.ALL, 5 )
		self.m_buttonMirror.SetToolTipString(u'�򿪾���ת����')
		
		self.m_buttonInvert = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ƬЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_2.Add( self.m_buttonInvert, 0, wx.ALL, 5 )
		self.m_buttonInvert.SetToolTipString(u'�򿪵�Ƭ����')
		
		self.m_buttonMolten = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_3.Add( self.m_buttonMolten, 0, wx.ALL, 5 )
		# No config needed
		self.m_buttonLighting = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�ƹ�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_3.Add( self.m_buttonLighting, 0, wx.ALL, 5 )
		self.m_buttonLighting.SetToolTipString(u'�򿪵ƹ�����')

		self.m_buttonSketch = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_3.Add( self.m_buttonSketch, 0, wx.ALL, 5 )
		self.m_buttonSketch.SetToolTipString(u'����������')
		
		self.m_buttonIce = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_3.Add( self.m_buttonIce, 0, wx.ALL, 5 )
		# No config needed
		
		self.m_buttonPapercut = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ֽЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_3.Add( self.m_buttonPapercut, 0, wx.ALL, 5 )
		self.m_buttonPapercut.SetToolTipString(u'�򿪼�ֽ����')
		
		self.m_buttonFocus = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�۹�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_3.Add( self.m_buttonFocus, 0, wx.ALL, 5 )
		self.m_buttonFocus.SetToolTipString(FOCUS_TIP)
		
		size = self.m_buttonPapercut.GetSize()
		
		self.m_buttonDeYellow = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ƭ", wx.DefaultPosition, size) #wx.DefaultSize) #, wx.BU_EXACTFIT )
		sbSizerTaskPanel_4.Add( self.m_buttonDeYellow, 0, wx.ALL, 5 )
		self.m_buttonDeYellow.SetToolTipString(u'������Ƭ����')
		
		self.m_buttonTan = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ƭ", wx.DefaultPosition, size) #wx.DefaultSize) #, wx.BU_EXACTFIT )
		sbSizerTaskPanel_4.Add( self.m_buttonTan, 0, wx.ALL, 5 )
		# No config needed
		
		self.m_buttonSwirl = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, size) #wx.DefaultSize) #, wx.BU_EXACTFIT )
		sbSizerTaskPanel_4.Add( self.m_buttonSwirl, 0, wx.ALL, 5 )
		self.m_buttonSwirl.SetToolTipString(u'����������')        
		
		self.m_buttonSpherize = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"������", wx.DefaultPosition, size) #wx.DefaultSize) #, wx.BU_EXACTFIT )
		sbSizerTaskPanel_4.Add( self.m_buttonSpherize, 0, wx.ALL, 5 )
		# No config needed
		# ƴͼ
		self.m_buttonPuzzle = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"ƴͼЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_4.Add( self.m_buttonPuzzle, 0, wx.ALL, 5 )
		self.m_buttonPuzzle.SetToolTipString(PUZZLE_TIP) 
		
		# �۽�
		self.m_buttonCornerFold = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�۽�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_4.Add( self.m_buttonCornerFold, 0, wx.ALL, 5 )
		self.m_buttonCornerFold.SetToolTipString(u'���۽�����')
		
		# �۽�
		self.m_buttonTear = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ݱ�Ե", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_5.Add( self.m_buttonTear, 0, wx.ALL, 5 )
		self.m_buttonTear.SetToolTipString(TEAR_TIP)
		
		self.m_buttonFade = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_5.Add( self.m_buttonFade, 0, wx.ALL, 5 )
		self.m_buttonFade.SetToolTipString(u'�򿪽�������')
		
		self.m_buttonMix = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ɫЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_5.Add( self.m_buttonMix, 0, wx.ALL, 5 )
		self.m_buttonMix.SetToolTipString(u'�򿪻�ɫ����')
		
		self.m_buttonReplace = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��ɫ�滻", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_5.Add( self.m_buttonReplace, 0, wx.ALL, 5 )
		self.m_buttonReplace.SetToolTipString(REPLACE_TIP)
		
		self.m_buttonAutoBrightness = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�Զ�����", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_5.Add( self.m_buttonAutoBrightness, 0, wx.ALL, 5 )
		self.m_buttonAutoBrightness.SetToolTipString(u'���Զ���������\r\n����ͼƬ��ָ��������')
		
		self.m_buttonRoll = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"����Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_5.Add( self.m_buttonRoll, 0, wx.ALL, 5 )
		self.m_buttonRoll.SetToolTipString(u'�򿪹�������')
		
		self.m_buttonShear = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"��бЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_6.Add( self.m_buttonShear, 0, wx.ALL, 5 )
		self.m_buttonShear.SetToolTipString(u'����б����')
		
		self.m_buttonMosaic = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"������", wx.DefaultPosition, self.m_buttonShear.GetSize(), 0)
		sbSizerTaskPanel_6.Add( self.m_buttonMosaic, 0, wx.ALL, 5 )
		self.m_buttonMosaic.SetToolTipString(u'������������')
		
		self.m_buttonMask = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�ɰ�Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_6.Add( self.m_buttonMask, 0, wx.ALL, 5 )
		self.m_buttonMask.SetToolTipString(u'���ɰ�����')
		
		self.m_buttonBlend = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"���Ч��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_6.Add( self.m_buttonBlend, 0, wx.ALL, 5 )
		self.m_buttonBlend.SetToolTipString(u'�򿪻������')
		
		self.m_buttonWaterMark = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"ˮӡЧ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_6.Add( self.m_buttonWaterMark, 0, wx.ALL, 5 )
		self.m_buttonWaterMark.SetToolTipString(u'��ˮӡ����')
		
		# �ָ�
		self.m_buttonSplit = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"ͼƬ�ָ�", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_6.Add( self.m_buttonSplit, 0, wx.ALL, 5 )
		self.m_buttonSplit.SetToolTipString(SPLIT_TIP)
		
		self.m_buttonSlice = wx.ToggleButton( self.m_panel1, wx.ID_ANY, u"�м�ȥ��", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )
		sbSizerTaskPanel_7.Add( self.m_buttonSlice, 0, wx.ALL, 5 )
		self.m_buttonSlice.SetToolTipString(SLICE_TIP) 
		
		self.m_bpButtonPlus = wx.Button( self.m_panel1, wx.ID_ANY, "+", wx.DefaultPosition, wx.DefaultSize, wx.BU_EXACTFIT )		
		sbSizerTaskPanel_7.Add( self.m_bpButtonPlus, 0, wx.ALL, 5 )
		
		self.m_bpButtonPlus.SetToolTipString(u'��ӵ������б�')
		sbSizerTaskPanel.Add(sbSizerTaskPanel_1, 1, wx.ALL | wx.TOP, 0 )
		sbSizerTaskPanel.Add(sbSizerTaskPanel_2, 1, wx.ALL | wx.TOP, 0 )
		sbSizerTaskPanel.Add(sbSizerTaskPanel_3, 1, wx.ALL | wx.TOP, 0 )
		sbSizerTaskPanel.Add(sbSizerTaskPanel_4, 1, wx.ALL | wx.TOP, 0 )
		sbSizerTaskPanel.Add(sbSizerTaskPanel_5, 1, wx.ALL | wx.TOP, 0 )
		sbSizerTaskPanel.Add(sbSizerTaskPanel_6, 1, wx.ALL | wx.TOP, 0 )
		sbSizerTaskPanel.Add(sbSizerTaskPanel_7, 1, wx.ALL | wx.TOP, 0 )
		
		#Task dict
		self.taskListBtns = [self.m_buttonRound, self.m_buttonBackground, self.m_buttonBorder, self.m_buttonShadow,
					 self.m_buttonReflection, self.m_buttonGrayscale, self.m_buttonBlkWht, self.m_buttonEnhance,
					 self.m_buttonFilter, self.m_buttonRotate, self.m_buttonMirror, self.m_buttonInvert, self.m_buttonMolten,
					 self.m_buttonLighting, self.m_buttonSketch, self.m_buttonIce, self.m_buttonPapercut, self.m_buttonFocus,
					 self.m_buttonDeYellow, self.m_buttonTan, self.m_buttonSwirl, self.m_buttonSpherize,
					 self.m_buttonPuzzle, self.m_buttonCornerFold, self.m_buttonTear, self.m_buttonFade, self.m_buttonMix,
					 self.m_buttonReplace, self.m_buttonAutoBrightness, self.m_buttonRoll, self.m_buttonShear, self.m_buttonMosaic,
					 self.m_buttonMask, self.m_buttonBlend, self.m_buttonWaterMark, self.m_buttonSplit, self.m_buttonSlice ]
		
		bSizerMain.Add( sbSizerTaskPanel, 1, wx.EXPAND, 0 )
		#---------------------panel ends ---------------------      
		#------------setting for round corner starts-----------
		bSizerSetting = wx.BoxSizer( wx.VERTICAL )
		self.staticRound = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"Բ������" )
		self.sbSizerRound = wx.StaticBoxSizer( self.staticRound , wx.VERTICAL )
		
		bSizerRoundCkBox = wx.BoxSizer( wx.VERTICAL )
		
		self.m_checkBoxDefault = wx.CheckBox( self.m_panel1, wx.ID_ANY, u"ѡ��Ĭ��", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_checkBoxDefault.SetValue(True)
		bSizerRoundCkBox.Add( self.m_checkBoxDefault, 0, wx.ALL, 5 )            
		
		self.sbSizerRound.Add( bSizerRoundCkBox, 1, wx.EXPAND, 5 )              
		
		bSizerValue = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_staticTextRound = wx.StaticText( self.m_panel1, wx.ID_ANY, u"Բ�ǰ뾶", wx.DefaultPosition, wx.DefaultSize, 0 )
		#self.m_staticTextRound.Wrap( -1 )
		#self.m_staticTextRound.Enable( False )
		bSizerValue.Add( self.m_staticTextRound, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlNum = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"5", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 50, 0 )
		bSizerValue.Add( self.m_spinCtrlNum, 0, wx.ALIGN_CENTER|wx.ALL, 5 )		
		
		self.m_staticTextSize = wx.StaticText( self.m_panel1, wx.ID_ANY, u"% ͼƬ�ߴ�", wx.DefaultPosition, wx.DefaultSize, 0 )
		#self.m_staticTextSize.Wrap( -1 )
		bSizerValue.Add( self.m_staticTextSize, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.sbSizerRound.Add( bSizerValue, 1, wx.EXPAND, 5 )
		bSizerSetting.Add( self.sbSizerRound, 1, wx.EXPAND, 0 )
		#------------setting for round corner ends-----------
		#------------setting for backgroound setting starts-----------
		self.staticBackground = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��������" )
		self.sbSizerBg = wx.StaticBoxSizer( self.staticBackground, wx.HORIZONTAL )
		
		bSizerSource = wx.BoxSizer( wx.VERTICAL )               
		bSizerSrcColor = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_radioBtnCkColor = wx.RadioButton( self.m_panel1, wx.ID_ANY, u"��ɫ", wx.DefaultPosition, wx.DefaultSize, wx.RB_GROUP )
		bSizerSrcColor.Add( self.m_radioBtnCkColor, 0, wx.ALL, 5 )		
		
		self.m_bgColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,13 ), wx.CLRP_USE_TEXTCTRL) #wx.CLRP_DEFAULT_STYLE )
		bSizerSrcColor.Add( self.m_bgColourPicker, 0, wx.ALL, 5 )
		#self.cur_bgColor = wx.BLACK
		
		bSizerSrcColor.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )
		
		self.m_methodText = wx.StaticText( self.m_panel1, label = u"���ӷ�ʽ")
		bSizerSrcColor.Add( self.m_methodText, 0, wx.ALL, 5 )
		self.m_MethodList = [u'ƽ��', u'����', u'ƫ��']
		self.m_comboMethod = wx.ComboBox( self.m_panel1, value = self.m_MethodList[2], choices = self.m_MethodList, style = wx.CB_READONLY)
		bSizerSrcColor.Add(self.m_comboMethod, 0, wx.ALL|wx.RIGHT, 5 )
		self.m_comboMethod.Bind(wx.EVT_COMBOBOX_CLOSEUP, self.onMethodChanged)   
		
		bSizerSource.Add( bSizerSrcColor, 1, wx.EXPAND, 5 )
		
		bSizerSrcPic = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_radioBtnCkPic = wx.RadioButton( self.m_panel1, wx.ID_ANY, u"ͼƬ", wx.DefaultPosition, wx.DefaultSize, 0 )
		bSizerSrcPic.Add( self.m_radioBtnCkPic, 0, wx.ALIGN_TOP|wx.ALL, 5 )
		
		self.m_textCtrlPath = wx.TextCtrl( self.m_panel1, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, (150, -1), wx.TE_READONLY)
		bSizerSrcPic.Add( self.m_textCtrlPath, 0, wx.ALL, 5 )
		self.pic = None         
		
		self.m_btnBrowse = wx.Button( self.m_panel1, wx.ID_ANY, u"...", wx.DefaultPosition, (-1, 21), wx.BU_EXACTFIT )
		bSizerSrcPic.Add( self.m_btnBrowse, 0, wx.ALL, 5 )              
		
		bSizerSource.Add( bSizerSrcPic, 1, wx.EXPAND, 5 )
		self.sbSizerBg.Add( bSizerSource, 1, wx.EXPAND, 5 )
		
		bSizerOffset = wx.BoxSizer( wx.VERTICAL )               
		bSizerOffsetH = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_staticTextHPos = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ˮƽλ��", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextHPos.Wrap( -1 )
		bSizerOffsetH.Add( self.m_staticTextHPos, 0, wx.ALL|wx.RIGHT, 5 )
		
		self.m_HPosList = [u'��', u'��', u'��']
		self.m_VPosList = [u'��', u'��', u'��']
		
		self.m_comboTextHPos = wx.ComboBox( self.m_panel1, value = self.m_HPosList[0], choices = self.m_HPosList, style = wx.CB_READONLY)
		bSizerOffsetH.Add( self.m_comboTextHPos, 0, wx.ALL|wx.RIGHT, 5 )
		
		#---
		self.m_staticTextHtxt = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ˮƽƫ��(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextHtxt.Wrap( -1 )
		bSizerOffsetH.Add( self.m_staticTextHtxt, 0, wx.ALL|wx.RIGHT, 5 )
		
		self.m_textCtrlHbox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '0', wx.DefaultPosition, (60, -1), 0 )
		bSizerOffsetH.Add( self.m_textCtrlHbox, 0, wx.ALL, 5 )
		#---
		
		bSizerOffset.Add( bSizerOffsetH, 1, wx.EXPAND, 5 )              
		bSizerV = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_staticTextVPos = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ֱλ��", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextVPos.Wrap( -1 )
		bSizerV.Add( self.m_staticTextVPos, 0, wx.ALL|wx.RIGHT, 5 )
		
		self.m_comboTextVPos = wx.ComboBox( self.m_panel1, value = self.m_VPosList[0], choices = self.m_VPosList, style = wx.CB_READONLY)
		bSizerV.Add( self.m_comboTextVPos, 0, wx.ALL|wx.RIGHT, 5 )
		
		#--------
		self.m_staticTextVtxt = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ֱƫ��(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextVtxt.Wrap( -1 )
		bSizerV.Add( self.m_staticTextVtxt, 0, wx.ALL|wx.RIGHT, 5 )
		
		self.m_textCtrlVbox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '0', wx.DefaultPosition, (60, -1), 0 )
		bSizerV.Add( self.m_textCtrlVbox, 0, wx.ALL, 5 )
		#----------             
		
		bSizerOffset.Add( bSizerV, 1, wx.EXPAND, 5 )
		self.sbSizerBg.Add( bSizerOffset, 1, wx.EXPAND, 5 )             
		
		bSizerSetting.Add( self.sbSizerBg, 1, wx.EXPAND, 0 )
		#-------------- Add border feature ------------     
		self.staticBorder = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�߿�����" )
		sbSizerStaticBorder = wx.StaticBoxSizer( self.staticBorder, wx.HORIZONTAL )
		
		self.m_staticTextBorder = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�߾�(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextBorder.Wrap( -1 )
		sbSizerStaticBorder.Add( self.m_staticTextBorder, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_textCtrlBorderBox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '20', wx.DefaultPosition, wx.Size( 80,-1 ), 0 )
		sbSizerStaticBorder.Add( self.m_textCtrlBorderBox, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		
		sbSizerStaticBorder.Add((20, -1))
		
		self.m_staticTextColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�߿���ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextColor.Wrap( -1 )
		sbSizerStaticBorder.Add( self.m_staticTextColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.cur_borderColor = wx.Colour(178, 178, 178)
		self.m_borderColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, self.cur_borderColor, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE )
		sbSizerStaticBorder.Add( self.m_borderColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )		
				
		sbSizerStaticBorder.Add((20, -1))
		
		self.m_staticTextOpacity = wx.StaticText( self.m_panel1, wx.ID_ANY, u'��͸����', wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextOpacity.Wrap( -1 )
		sbSizerStaticBorder.Add( self.m_staticTextOpacity, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_sliderOpacity = wx.Slider( self.m_panel1, wx.ID_ANY, 80, 0, 100, wx.DefaultPosition, wx.DefaultSize, style=wx.SL_HORIZONTAL|wx.SL_LABELS )
		sbSizerStaticBorder.Add( self.m_sliderOpacity, 0, wx.ALIGN_CENTER|wx.ALL, 5 )     

		bSizerSetting.Add(sbSizerStaticBorder, 1, wx.EXPAND, 0)
		#---------------border feature ends -----------
		#---------------shadow feature starts ----------
		self.staticShadow = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��Ӱ����" )
		sbSizerStaticShadow = wx.StaticBoxSizer( self.staticShadow, wx.HORIZONTAL )
		
		#sbSizerStaticShadow.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )
		
		self.m_staticHTextShadow = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ˮƽƫ��(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticHTextShadow.Wrap( -1 )
		sbSizerStaticShadow.Add( self.m_staticHTextShadow, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlShadowHBox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '10', wx.DefaultPosition, wx.Size( 80,-1 ), 0 )
		sbSizerStaticShadow.Add( self.m_textCtrlShadowHBox, 0, wx.ALIGN_CENTER, 5 )
		
		sbSizerStaticShadow.Add( (20, -1))
		
		self.m_staticVTextShadow = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ֱƫ��(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticVTextShadow.Wrap( -1 )
		sbSizerStaticShadow.Add( self.m_staticVTextShadow, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlShadowVBox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '10', wx.DefaultPosition, wx.Size( 80,-1 ), 0 )
		sbSizerStaticShadow.Add( self.m_textCtrlShadowVBox, 0, wx.ALIGN_CENTER, 5 )      

		bSizerSetting.Add( sbSizerStaticShadow, 1, wx.EXPAND, 0 )       
		#---------------shadow feature ends ------------
		#--------------- reflection starts -------------   
		self.staticReflection = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��Ӱ����" )
		sbSizerStaticReflection = wx.StaticBoxSizer( self.staticReflection, wx.HORIZONTAL )
		
		self.m_staticTextDepth = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��Ӱ���(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextDepth.Wrap( -1 )
		sbSizerStaticReflection.Add( self.m_staticTextDepth, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlDepthBox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '500', wx.DefaultPosition, wx.Size( 80,-1 ), 0 )
		sbSizerStaticReflection.Add( self.m_textCtrlDepthBox, 0, wx.ALIGN_CENTER, 5 )
		
		sbSizerStaticReflection.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )
		
		self.m_staticReflectionBgColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticReflectionBgColor.Wrap( -1 )
		sbSizerStaticReflection.Add( self.m_staticReflectionBgColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_ReflectionBgColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE )
		sbSizerStaticReflection.Add( self.m_ReflectionBgColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		#self.cur_ReflectionBgColor = wx.BLACK
		
		sbSizerStaticReflection.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )
		
		self.m_staticReflectionOpacity = wx.StaticText( self.m_panel1, wx.ID_ANY, u'��͸����', wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticReflectionOpacity.Wrap( -1 )
		sbSizerStaticReflection.Add( self.m_staticReflectionOpacity, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_sliderReflectionOpacity = wx.Slider( self.m_panel1, wx.ID_ANY, 0, 0, 100, wx.DefaultPosition, wx.DefaultSize, style=wx.SL_HORIZONTAL|wx.SL_LABELS )
		sbSizerStaticReflection.Add( self.m_sliderReflectionOpacity, 0, wx.ALIGN_CENTER | wx.LEFT, 5 )

		bSizerSetting.Add( sbSizerStaticReflection, 1, wx.EXPAND, 0 )
		#----------------reflection ends ---------------
		#---------------enhance starts ------------------
		self.staticEnhance = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"ͼ����ǿ����" )
		sbSizerStaticEnhance = wx.StaticBoxSizer( self.staticEnhance, wx.HORIZONTAL )
		#sbSizerStaticEnhance.Add(self.staticEnhance, 0, wx.ALL|wx.RIGHT, 5 )
		
		self.m_staticTextFactor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ǿϵ��(-100 ~ 100)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextFactor.Wrap( -1 )
		sbSizerStaticEnhance.Add( self.m_staticTextFactor, 0, wx.ALIGN_CENTER | wx.RIGHT, 5 )    
		
		self.m_spinCtrlFactor = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"20", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, -100, 100, 20 )
		sbSizerStaticEnhance.Add( self.m_spinCtrlFactor, 0, wx.ALIGN_CENTER| wx.RIGHT, 5 )
		
		sbSizerStaticEnhance.Add( ( 20, 0) )
		
		self.m_staticTextEnhance = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ǿ����", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextEnhance.Wrap( -1 )
		sbSizerStaticEnhance.Add( self.m_staticTextEnhance, 0, wx.ALIGN_CENTER| wx.RIGHT, 5 )
				
		self.m_EnhanceModeList = [u'��', u'ɫ��', u'����', u'�Աȶ�']      
		self.m_EnhanceBox = wx.ComboBox( self.m_panel1, value = self.m_EnhanceModeList[3], choices = self.m_EnhanceModeList, style = wx.CB_READONLY)
		sbSizerStaticEnhance.Add( self.m_EnhanceBox, 0, wx.ALIGN_CENTER, 5 )     

		bSizerSetting.Add( sbSizerStaticEnhance, 1, wx.EXPAND, 0 )
		#---------------enhance ends --------------------
		#---------------Filter starts ------------------
		self.staticFilter = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�˾�����" )
		sbSizerStaticFilter = wx.StaticBoxSizer( self.staticFilter, wx.HORIZONTAL )
				
		#self.m_FilterModeList = [u'ģ��', u'��˹ģ��', u'����', u'����', u'��Ե��ǿ', u'ƽ��', u'��ֵ�˲�', u'��', u'ϸ��']
		self.m_checkBoxBlur = wx.CheckBox(self.m_panel1, label=u'ģ��', style = wx.BU_EXACTFIT )
		self.m_checkBoxGBlur = wx.CheckBox(self.m_panel1, label=u'��˹ģ��', style = wx.BU_EXACTFIT )
		self.m_checkBoxEmboss = wx.CheckBox(self.m_panel1, label=u'����', style = wx.BU_EXACTFIT )
		self.m_checkBoxContour = wx.CheckBox(self.m_panel1, label=u'����', style = wx.BU_EXACTFIT )
		self.m_checkBoxEdge = wx.CheckBox(self.m_panel1, label=u'��Ե��ǿ', style = wx.BU_EXACTFIT )
		self.m_checkBoxSmooth = wx.CheckBox(self.m_panel1, label=u'ƽ��', style = wx.BU_EXACTFIT )
		self.m_checkBoxMedFilter = wx.CheckBox(self.m_panel1, label=u'��ֵ�˲�', style = wx.BU_EXACTFIT )
		self.m_checkBoxSharpen = wx.CheckBox(self.m_panel1, label=u'��', style = wx.BU_EXACTFIT )
		self.m_checkBoxDetail = wx.CheckBox(self.m_panel1, label=u'ϸ��', style = wx.BU_EXACTFIT )
		self.m_FilterBoxList = [self.m_checkBoxBlur, self.m_checkBoxGBlur, self.m_checkBoxEmboss,self.m_checkBoxContour,
					self.m_checkBoxEdge, self.m_checkBoxSmooth, self.m_checkBoxMedFilter, self.m_checkBoxSharpen,
					self.m_checkBoxDetail ]
		self.m_checkBoxGBlur.SetValue(True)
		for checkbox in self.m_FilterBoxList:           
			sbSizerStaticFilter.Add(checkbox, 0, wx.ALL|wx.CENTER, 3 )		
		
		bSizerSetting.Add( sbSizerStaticFilter, 1, wx.EXPAND, 0 )
		#---------------Filter ends ------------------
		#---------------rotate starts ------------------        
		self.staticRotate = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��ת����" )
		sbSizerStaticRotate = wx.StaticBoxSizer( self.staticRotate, wx.HORIZONTAL )		

		self.m_radioBtnFixed = wx.RadioButton( self.m_panel1, wx.ID_ANY, u"�̶��Ƕ�", wx.DefaultPosition, wx.DefaultSize, wx.RB_GROUP )		
		sbSizerStaticRotate.Add( self.m_radioBtnFixed, 0, wx.ALIGN_CENTER | wx.RIGHT, 5 )		
		
		m_choiceFixedChoices = [ u"90", u"180", u"270" ]
		self.m_choiceFixed = wx.Choice( self.m_panel1, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, m_choiceFixedChoices, 0 )
		self.m_choiceFixed.SetSelection( 0 )
		sbSizerStaticRotate.Add( self.m_choiceFixed, 0, wx.ALIGN_CENTER| wx.RIGHT, 5 )		
		
		sbSizerStaticRotate.Add( ( 20, -1))
		
		self.m_radioBtnAny = wx.RadioButton( self.m_panel1, wx.ID_ANY, u"����Ƕ�(0~360)", wx.DefaultPosition, wx.DefaultSize, 0 )		
		sbSizerStaticRotate.Add( self.m_radioBtnAny, 0, wx.ALIGN_CENTER| wx.RIGHT, 5 )
		
		self.m_spinCtrlAny = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size(80, -1), wx.SP_ARROW_KEYS, 0, 360, 0 )
		self.m_spinCtrlAny.Enable(False)		
		sbSizerStaticRotate.Add( self.m_spinCtrlAny, 0, wx.ALIGN_CENTER, 5 )			

		bSizerSetting.Add( sbSizerStaticRotate, 1, wx.EXPAND, 0 )
		#---------------rotate ends --------------------
		#---------------mirror starts ------------------        
		self.staticMirror = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"����ת����" )
		sbSizerStaticMirror = wx.StaticBoxSizer( self.staticMirror, wx.HORIZONTAL )
		
		self.m_staticTextMirror = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ת����", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextMirror.Wrap( -1 )
		sbSizerStaticMirror.Add( self.m_staticTextMirror, 0, wx.ALIGN_CENTER | wx.RIGHT, 5 )  
		
		self.m_MirrorModeList = [u'���¾���', u'���Ҿ���', u'���߽�ѡ']     
		self.m_MirrorBox = wx.ComboBox( self.m_panel1, value = self.m_MirrorModeList[0], choices = self.m_MirrorModeList, style = wx.CB_READONLY)
		sbSizerStaticMirror.Add( self.m_MirrorBox, 0, wx.ALIGN_CENTER, 5 )  

		bSizerSetting.Add( sbSizerStaticMirror, 1, wx.EXPAND, 0 )
		#---------------mirror ends --------------------
		#--------------- invert starts -----------------
		self.staticInvert = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��Ƭ����" )
		sbSizerStaticInvert = wx.StaticBoxSizer( self.staticInvert, wx.HORIZONTAL )
		
		self.m_staticTextInvert = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��������(-50 �� 50)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextInvert.Wrap( -1 )
		sbSizerStaticInvert.Add( self.m_staticTextInvert, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_spinCtrlInvert = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"50", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, -50, 50, 50 )
		sbSizerStaticInvert.Add( self.m_spinCtrlInvert, 0, wx.ALIGN_CENTER, 5 )
		
		bSizerSetting.Add( sbSizerStaticInvert, 1, wx.EXPAND, 0 )
		#----------------invert ends -------------------
		#---------------- lighting starts -------------------
		self.staticLighting = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�ƹ�����" )
		sbSizerStaticLighting = wx.StaticBoxSizer( self.staticLighting, wx.HORIZONTAL )
		
		self.m_staticTextLightingPos = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��Դλ��", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextLightingPos.Wrap( -1 )
		sbSizerStaticLighting.Add( self.m_staticTextLightingPos, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_PosList = [u'����', u'����', u'����', u'����', u'����', u'����', u'����', u'����',u'����']	
		self.m_comboTextLightingPos = wx.ComboBox( self.m_panel1, value = self.m_PosList[4], choices = self.m_PosList, style = wx.CB_READONLY)
		sbSizerStaticLighting.Add( self.m_comboTextLightingPos, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		sbSizerStaticLighting.Add( ( 20, -1))
		
		self.m_staticTextLighting = wx.StaticText( self.m_panel1, wx.ID_ANY, u"����ǿ��(0 �� 50)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextLighting.Wrap( -1 )
		sbSizerStaticLighting.Add( self.m_staticTextLighting, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_spinCtrlLighting = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"20", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 50, 20 )
		sbSizerStaticLighting.Add( self.m_spinCtrlLighting, 0, wx.ALIGN_CENTER, 5 )
		
		bSizerSetting.Add( sbSizerStaticLighting, 1, wx.EXPAND, 0 )
		#----------------lighting ends -------------------		
		#---------------- sketch starts -------------------
		self.staticSketch = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��������" )
		sbSizerStaticSketch = wx.StaticBoxSizer( self.staticSketch, wx.HORIZONTAL )
		
		self.m_staticTextSketch = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�߽�ֵ(0 ��100)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextSketch.Wrap( -1 )
		sbSizerStaticSketch.Add( self.m_staticTextSketch, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_spinCtrlSketch = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"15", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 15 )
		sbSizerStaticSketch.Add( self.m_spinCtrlSketch, 0, wx.ALIGN_CENTER, 5 )
		
		bSizerSetting.Add( sbSizerStaticSketch, 1, wx.EXPAND, 0 )
		#----------------sketch ends -------------------		
		#---------------- papercut starts -------------------
		self.staticPapercut = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��ֽ����" )
		sbSizerStaticPapercut = wx.StaticBoxSizer( self.staticPapercut, wx.HORIZONTAL )
		
		self.m_staticPapercutBgColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticPapercutBgColor.Wrap( -1 )
		sbSizerStaticPapercut.Add( self.m_staticPapercutBgColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_PapercutBgColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE )
		sbSizerStaticPapercut.Add( self.m_PapercutBgColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.cur_PapercutBgColor = wx.BLACK
		
		sbSizerStaticPapercut.Add( ( 20, -1))
		
		self.m_staticPapercutFgColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ǰ����ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticPapercutFgColor.Wrap( -1 )		
		sbSizerStaticPapercut.Add( self.m_staticPapercutFgColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_PapercutFgColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.RED, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE )
		sbSizerStaticPapercut.Add( self.m_PapercutFgColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.cur_PapercutFgColor = wx.RED
		
		sbSizerStaticPapercut.Add( ( 20, -1))
		
		self.m_staticTextPapercut = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�߽�ֵ(0 ��255)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextPapercut.Wrap( -1 )
		sbSizerStaticPapercut.Add( self.m_staticTextPapercut, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_spinCtrlPapercut = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"20", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 255, 20 )
		sbSizerStaticPapercut.Add( self.m_spinCtrlPapercut, 0, wx.ALIGN_CENTER, 5 )
		
		bSizerSetting.Add( sbSizerStaticPapercut, 1, wx.EXPAND, 0 )
		#----------------papercut ends -------------------
		#---------------- deYellow starts -------------------
		self.staticDeYellow = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"����Ƭ����" )
		sbSizerStaticDeYellow = wx.StaticBoxSizer( self.staticDeYellow, wx.HORIZONTAL )
		
		self.m_radioBoxDeYellow_1 = wx.RadioButton( self.m_panel1, wx.ID_ANY, u"����1", wx.DefaultPosition, wx.DefaultSize, wx.RB_GROUP )
		self.m_radioBoxDeYellow_2 = wx.RadioButton( self.m_panel1, wx.ID_ANY, u"����2", wx.DefaultPosition, wx.DefaultSize, 0 )		
		sbSizerStaticDeYellow.Add(self.m_radioBoxDeYellow_1, 0, wx.ALL|wx.CENTER, 5 )
		sbSizerStaticDeYellow.Add(self.m_radioBoxDeYellow_2, 0, wx.ALL|wx.CENTER, 5 )		
		bSizerSetting.Add( sbSizerStaticDeYellow, 1, wx.EXPAND, 0 )		
		#----------------deYellow ends -------------------
		#---------------- swirl starts -------------------
		self.staticSwirl = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��������" )
		sbSizerStaticSwirl = wx.StaticBoxSizer( self.staticSwirl, wx.HORIZONTAL )
		
		self.m_staticTextSwirl = wx.StaticText( self.m_panel1, wx.ID_ANY, u"���д�С", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextSwirl.Wrap( -1 )
		sbSizerStaticSwirl.Add( self.m_staticTextSwirl, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_sliderSwirlOpacity = wx.Slider( self.m_panel1, wx.ID_ANY, 50, 0, 100, wx.DefaultPosition, wx.DefaultSize, style=wx.SL_HORIZONTAL|wx.SL_LABELS )
		sbSizerStaticSwirl.Add( self.m_sliderSwirlOpacity, 0, wx.ALIGN_CENTER | wx.LEFT, 5 )
		
		bSizerSetting.Add( sbSizerStaticSwirl, 1, wx.EXPAND, 0 )
		#----------------swirl ends -------------------
		#--------------- puzzle starts ------------------
		self.staticPuzzle = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"ƴͼ����" )
		sbSizerStaticPuzzle = wx.StaticBoxSizer( self.staticPuzzle, wx.VERTICAL )
		bSizerPuzzleUp = wx.BoxSizer( wx.HORIZONTAL )
		bSizerPuzzleDown = wx.BoxSizer( wx.HORIZONTAL )
		
		self.m_staticTextQtyX = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ˮƽ��Ŀ", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextQtyX.Wrap( -1 )
		bSizerPuzzleUp.Add( self.m_staticTextQtyX, 0, wx.ALL, 5 )    
		
		self.m_spinCtrlQtyX = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"2", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 1, 100, 2 )
		bSizerPuzzleUp.Add( self.m_spinCtrlQtyX, 0, wx.ALL, 5 )
		
		bSizerPuzzleUp.Add( ( 5, -1) )
		
		self.m_staticTextQtyY = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ֱ��Ŀ", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextQtyY.Wrap( -1 )
		bSizerPuzzleUp.Add( self.m_staticTextQtyY, 0, wx.ALL, 5 )    
		
		self.m_spinCtrlQtyY = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"2", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 1, 100, 2 )
		bSizerPuzzleUp.Add( self.m_spinCtrlQtyY, 0, wx.ALL, 5 )
		
		bSizerPuzzleUp.Add( ( 5, -1) )
		
		self.m_checkBoxResize = wx.CheckBox(self.m_panel1, label=u'��������', style = wx.BU_EXACTFIT )
		bSizerPuzzleUp.Add( self.m_checkBoxResize, 0, wx.ALL, 5 )
				
		bSizerPuzzleUp.Add( ( 5, -1) )
		
		self.m_staticPuzzleBgColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		bSizerPuzzleUp.Add( self.m_staticPuzzleBgColor, 0, wx.ALL, 5 )
		
		self.m_PuzzleBgColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.WHITE, wx.DefaultPosition, wx.Size( 45,12 ), wx.CLRP_DEFAULT_STYLE )
		bSizerPuzzleUp.Add( self.m_PuzzleBgColourPicker, 0, wx.ALL, 5 )
		
		bSizerPuzzleUp.Add( ( 5, -1) )	
		
		self.m_PuzzleOrderList = [u'����', u'����']      
		self.m_PuzzleOrderBox = wx.ComboBox( self.m_panel1, value = self.m_PuzzleOrderList[0], choices = self.m_PuzzleOrderList, style = wx.CB_READONLY)
		bSizerPuzzleDown.Add( self.m_PuzzleOrderBox, 0, wx.ALL, 5 )
		
		bSizerPuzzleDown.Add( ( 5, -1) )
		
		#self.m_staticTextSrc = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ͼƬ��Դ", wx.DefaultPosition, wx.DefaultSize, 0 )
		#self.m_staticTextSrc.Wrap( -1 )
		#bSizerPuzzleDown.Add( self.m_staticTextSrc, 0, wx.ALIGN_CENTER |wx.RIGHT, 5 )
				
		self.m_PuzzleSrcModeList = [u'�ظ���ǰͼƬ', u'���м���ͼƬ']      
		self.m_PuzzleSrcBox = wx.ComboBox( self.m_panel1, value = self.m_PuzzleSrcModeList[0], choices = self.m_PuzzleSrcModeList, style = wx.CB_READONLY)
		bSizerPuzzleDown.Add( self.m_PuzzleSrcBox, 0, wx.ALL, 5 )
		
		bSizerPuzzleDown.Add( ( 15, -1) )
		
		self.m_staticTextGap = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ͼƬ���(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticTextGap.Wrap( -1 )
		bSizerPuzzleDown.Add( self.m_staticTextGap, 0, wx.ALL, 5 )    
		
		self.m_spinCtrlGap = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 0 )
		bSizerPuzzleDown.Add( self.m_spinCtrlGap, 0, wx.ALL, 5 )
		
		self.m_buttonListView = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.file_list, wx.DefaultPosition, wx.Size(24, 24), wx.NO_BORDER |wx.BU_EXACTFIT  )
		bSizerPuzzleDown.Add( self.m_buttonListView, 0, wx.ALL, 5 )
		self.m_buttonListView.SetToolTipString(u'�鿴�ļ��б�')
		
		#self.m_buttonListView = wx.Button( self.m_panel1, wx.ID_ANY, u"�鿴�б�", wx.DefaultPosition, wx.DefaultSize, 0 )
		#bSizerPuzzleDown.Add( self.m_buttonListView, 0, wx.ALL, 5 )
		
		sbSizerStaticPuzzle.Add(bSizerPuzzleUp, 1, wx.EXPAND, 5)
		sbSizerStaticPuzzle.Add(bSizerPuzzleDown, 1, wx.EXPAND, 5)

		#self.cur_PuzzleBgColor = (255,255,255,255) #(0, 255, 255, 255)

		bSizerSetting.Add( sbSizerStaticPuzzle, 1, wx.EXPAND, 5 )
		
		#bSizerMain.Add( bSizerSetting, 1, wx.EXPAND, 5 )
		#--------------- puzzle ends -------------------------------
		#--------------- flash light starts ------------------------
		self.staticFocus = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�۹�����" )
		sbSizerStaticFocus = wx.StaticBoxSizer( self.staticFocus, wx.HORIZONTAL )
		
		self.m_staticTextRadiusIn = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��Ȧ�뾶(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticFocus.Add( self.m_staticTextRadiusIn, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlRadiusIn = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"60", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 200, 60 )
		sbSizerStaticFocus.Add( self.m_spinCtrlRadiusIn, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticFocus.Add((10, -1))
		
		self.m_staticTextRadiusOut = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��Ȧ�뾶(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticFocus.Add( self.m_staticTextRadiusOut, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlRadiusOut = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"100", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 400, 100 )
		sbSizerStaticFocus.Add( self.m_spinCtrlRadiusOut, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticFocus.Add((10, -1))
		
		self.m_staticFocusEdgeColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��Ե��ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticFocus.Add( self.m_staticFocusEdgeColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_FocusEdgeColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE )
		sbSizerStaticFocus.Add( self.m_FocusEdgeColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		sbSizerStaticFocus.Add((10, -1))
		
		self.m_checkBoxFocusEdge = wx.CheckBox(self.m_panel1, label=u'��Ե͸��', style = wx.BU_EXACTFIT )
		sbSizerStaticFocus.Add( self.m_checkBoxFocusEdge, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		bSizerSetting.Add( sbSizerStaticFocus, 1, wx.EXPAND, 0 )
		#---------------- flash light ends -------------------------
		
		#--------------- edge fold starts ------------------------
		self.staticCornerFold = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�۽�����" )
		sbSizerStaticCornerFold = wx.StaticBoxSizer( self.staticCornerFold, wx.HORIZONTAL )
		
		self.m_staticTextEdge = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�߾�(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticCornerFold.Add( self.m_staticTextEdge, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlEdge = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"15", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 15 )
		sbSizerStaticCornerFold.Add( self.m_spinCtrlEdge, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticCornerFold.Add((10, -1))
		
		self.m_staticTextWhitePercent = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ɫ��ǿ(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticCornerFold.Add( self.m_staticTextWhitePercent, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlWhitePercent = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"80", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 80 )
		sbSizerStaticCornerFold.Add( self.m_spinCtrlWhitePercent, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticCornerFold.Add((10, -1))
		
		self.m_staticFoldPos = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�۽�λ��", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticCornerFold.Add( self.m_staticFoldPos, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_FoldPosModeList = [u'���', u'�ұ�']      
		self.m_FoldPosBox = wx.ComboBox( self.m_panel1, value = self.m_FoldPosModeList[1], choices = self.m_FoldPosModeList, style = wx.CB_READONLY)
		sbSizerStaticCornerFold.Add( self.m_FoldPosBox, 0, wx.ALIGN_CENTER, 5 )
		bSizerSetting.Add( sbSizerStaticCornerFold, 1, wx.EXPAND, 0 )

		#---------------- edge fold ends -------------------------
		#--------------- tear starts ------------------------
		self.staticTear = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��ݱ�Ե����" )
		sbSizerStaticTear = wx.StaticBoxSizer( self.staticTear, wx.HORIZONTAL )		
		
		self.m_staticTearDirection = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ݷ���", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticTear.Add( self.m_staticTearDirection, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_TearPosModeList = [u'����', u'����']      
		self.m_TearPosBox = wx.ComboBox( self.m_panel1, value = self.m_TearPosModeList[0], choices = self.m_TearPosModeList, style = wx.CB_READONLY)
		sbSizerStaticTear.Add( self.m_TearPosBox, 0, wx.ALIGN_CENTER, 5 )
		sbSizerStaticTear.Add((10, -1))
		
		self.m_staticTextTear = wx.StaticText( self.m_panel1, wx.ID_ANY, u"���λ��(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticTear.Add( self.m_staticTextTear, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlTear = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"80", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 80 )
		sbSizerStaticTear.Add( self.m_spinCtrlTear, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_checkBoxIsCut = wx.CheckBox( self.m_panel1, wx.ID_ANY, u"�Զ��ü�", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_checkBoxIsCut.SetValue(True)
		sbSizerStaticTear.Add( self.m_checkBoxIsCut, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		bSizerSetting.Add( sbSizerStaticTear, 1, wx.EXPAND, 0 )
		#--------------- tear ends ------------------------
		#--------------- Fade starts ------------------------
		self.staticFade = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��������" )
		sbSizerStaticFade = wx.StaticBoxSizer( self.staticFade, wx.HORIZONTAL )
		
		self.m_staticTextFadePos = wx.StaticText( self.m_panel1, wx.ID_ANY, u"����λ��(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticFade.Add( self.m_staticTextFadePos, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlFadePos = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"70", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 70 )
		sbSizerStaticFade.Add( self.m_spinCtrlFadePos, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticFade.Add((10, -1))
		
		self.m_staticFadeMode = wx.StaticText( self.m_panel1, wx.ID_ANY, u"����ģʽ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticFade.Add( self.m_staticFadeMode, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_FadeModeList = [u'����', u'����']      
		self.m_FadeModeBox = wx.ComboBox( self.m_panel1, value = self.m_FadeModeList[1], choices = self.m_FadeModeList, style = wx.CB_READONLY)
		sbSizerStaticFade.Add( self.m_FadeModeBox, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		sbSizerStaticFade.Add((10, -1))
		
		self.m_staticFadeColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticFade.Add( self.m_staticFadeColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_FadeColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE )
		sbSizerStaticFade.Add( self.m_FadeColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		bSizerSetting.Add( sbSizerStaticFade, 1, wx.EXPAND, 0 )
		
		#--------------- Mix starts ------------------------
		self.staticMix = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��ɫ����" )
		sbSizerStaticMix = wx.StaticBoxSizer( self.staticMix, wx.HORIZONTAL )
		
		self.m_staticTextMixFactor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ǿϵ��(%)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticMix.Add( self.m_staticTextMixFactor, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlMixFactor = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"20", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 20 )
		sbSizerStaticMix.Add( self.m_spinCtrlMixFactor, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticMix.Add((10, -1))
		
		self.m_staticMixColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�����ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticMix.Add( self.m_staticMixColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )

		self.m_MixColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		sbSizerStaticMix.Add(self.m_MixColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5)
				
		self.ctrlColor = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '(0, 0, 0, 255)', wx.DefaultPosition, wx.Size( 120,-1 ), wx.TE_PROCESS_ENTER )
		sbSizerStaticMix.Add( self.ctrlColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.ctrlColor.SetToolTipString(COLOR_FORMAT_TIP)
		self.ctrlColor.Bind(wx.EVT_TEXT, self.OnMixTxtChanged)
		
		bSizerSetting.Add( sbSizerStaticMix, 1, wx.EXPAND, 0 )
		
				
		#--------------- color replace starts ------------------------
		self.staticReplace = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��ɫ�滻����" )
		sbSizerStaticReplace = wx.StaticBoxSizer( self.staticReplace, wx.VERTICAL )		
		sbSizerStaticRegion = wx.BoxSizer(wx.HORIZONTAL )
		sbSizerStaticColorReplace = wx.BoxSizer(wx.HORIZONTAL )
		
		self.replaceModeStyleList = [u'��ɫ�滻', u'��ɫ���']
		self.m_checkBoxReplaceStyle = wx.ComboBox(self.m_panel1, value = self.replaceModeStyleList[0], choices = self.replaceModeStyleList, style = wx.CB_READONLY )
		sbSizerStaticRegion.Add( self.m_checkBoxReplaceStyle, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.m_staticTextModeTip = wx.StaticText( self.m_panel1, label = u"�������ѡ��������Ĭ��������ͼƬ��" )
		self.m_staticTextModeTip.Enable( False )
		sbSizerStaticRegion.Add( self.m_staticTextModeTip, 0, wx.ALIGN_CENTER|wx.LEFT, 5 ) #wx.LEFT | wx.TOP, 22 )
		self.m_checkBoxReplaceStyle.Bind(wx.EVT_COMBOBOX_CLOSEUP, self.OnReplaceModeChanged)
		
		self.m_staticOldColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"����ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticColorReplace.Add( self.m_staticOldColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		fgSizerReplace = wx.FlexGridSizer( 2, 3, 0, 0 )
		fgSizerReplace.SetFlexibleDirection( wx.HORIZONTAL )
		fgSizerReplace.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
		
		self.m_ColorCompare = [u'>',u'=', u'��',u'<',u'��']      
		self.m_ColorCompareBox = wx.ComboBox( self.m_panel1, value = self.m_ColorCompare[1], choices = self.m_ColorCompare, style = wx.CB_READONLY)
		fgSizerReplace.Add( self.m_ColorCompareBox, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.m_ColorCompareBox.Bind (wx.EVT_COMBOBOX, self.OnColorComboChanged )

		self.m_OldColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.BLACK, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		fgSizerReplace.Add(self.m_OldColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5)	
				
		self.ctrlOldColor = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '(0, 0, 0, 255)', wx.DefaultPosition, wx.Size( 120,-1 ), wx.TE_PROCESS_ENTER )
		fgSizerReplace.Add( self.ctrlOldColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
				
		self.ctrlOldColor.SetToolTipString(COLOR_FORMAT_TIP)
		self.ctrlOldColor.Bind(wx.EVT_TEXT, self.OnOldTxtChanged)
		
		self.m_staticTextOldMax = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��", wx.DefaultPosition, wx.DefaultSize, 0 )
		fgSizerReplace.Add( self.m_staticTextOldMax, 0, wx.ALIGN_RIGHT|wx.ALL, 5 )
		
		self.m_OldColourMaxPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.WHITE, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		fgSizerReplace.Add(self.m_OldColourMaxPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5)	
				
		self.ctrlOldColorMax = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '(255, 255, 255, 255)', wx.DefaultPosition, wx.Size( 120,-1 ), wx.TE_PROCESS_ENTER )
		fgSizerReplace.Add( self.ctrlOldColorMax, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
				
		self.ctrlOldColorMax.SetToolTipString(COLOR_FORMAT_TIP)
		self.ctrlOldColorMax.Bind(wx.EVT_TEXT, self.OnOldTxtMaxChanged)
		
		sbSizerStaticColorReplace.Add( fgSizerReplace, 1, wx.ALIGN_CENTER|wx.LEFT, 0 )
		
		sbSizerStaticReplace.Add((15, -1))
		
		self.m_staticNewColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"����ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticColorReplace.Add( self.m_staticNewColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )

		self.m_NewColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.RED, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		sbSizerStaticColorReplace.Add(self.m_NewColourPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5)
				
		self.ctrlNewColor = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '(255, 0, 0, 255)', wx.DefaultPosition, wx.Size( 120,-1 ), wx.TE_PROCESS_ENTER )
		sbSizerStaticColorReplace.Add( self.ctrlNewColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.ctrlNewColor.SetToolTipString(COLOR_FORMAT_TIP)
		self.ctrlNewColor.Bind(wx.EVT_TEXT, self.OnNewTxtChanged)
		
		sbSizerStaticReplace.Add( sbSizerStaticRegion, 1, wx.EXPAND, 0 )
		sbSizerStaticReplace.Add( sbSizerStaticColorReplace, 1, wx.EXPAND, 0 )
		bSizerSetting.Add( sbSizerStaticReplace, 1, wx.EXPAND, 0 )
		#--------------- Auto brightness starts ------------------------
		self.staticAutoBrightness = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�Զ���������" )
		sbSizerStaticAutoBrightness = wx.StaticBoxSizer( self.staticAutoBrightness, wx.HORIZONTAL )
		
		self.m_staticCurBrightness = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ǰ����", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticAutoBrightness.Add( self.m_staticCurBrightness, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.ctrlCurBrightness = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '', wx.DefaultPosition, wx.Size( 80,-1 ), wx.TE_READONLY )
		sbSizerStaticAutoBrightness.Add( self.ctrlCurBrightness, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		sbSizerStaticAutoBrightness.Add((15, -1))
		
		self.m_staticNewBrightness = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticAutoBrightness.Add( self.m_staticNewBrightness, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.m_spinCtrlNewBrightness = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"200", wx.DefaultPosition, wx.Size( 80,-1 ), wx.SP_ARROW_KEYS, 0, 255, 200 )
		sbSizerStaticAutoBrightness.Add( self.m_spinCtrlNewBrightness, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		bSizerSetting.Add( sbSizerStaticAutoBrightness, 1, wx.EXPAND, 0 )
		#--------------- roll starts --------------------------
		self.staticRoll = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��������" )
		sbSizerStaticRoll = wx.StaticBoxSizer( self.staticRoll, wx.HORIZONTAL )
		
		#sbSizerStaticRoll.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )
		
		self.m_staticHTextRoll = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ˮƽƫ��(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticHTextRoll.Wrap( -1 )
		sbSizerStaticRoll.Add( self.m_staticHTextRoll, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlRollHBox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '10', wx.DefaultPosition, wx.Size( 80,-1 ), 0 )
		sbSizerStaticRoll.Add( self.m_textCtrlRollHBox, 0, wx.ALIGN_CENTER, 5 )
		
		sbSizerStaticRoll.Add( (20, -1))
		
		self.m_staticVTextRoll = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ֱƫ��(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		self.m_staticVTextRoll.Wrap( -1 )
		sbSizerStaticRoll.Add( self.m_staticVTextRoll, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlRollVBox = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '10', wx.DefaultPosition, wx.Size( 80,-1 ), 0 )
		sbSizerStaticRoll.Add( self.m_textCtrlRollVBox, 0, wx.ALIGN_CENTER, 5 )      

		bSizerSetting.Add( sbSizerStaticRoll, 1, wx.EXPAND, 0 ) 
		#---------------  shear starts ------------------------
		self.staticShear = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"��бЧ������" )
		sbSizerStaticShear = wx.StaticBoxSizer( self.staticShear, wx.HORIZONTAL )
		
		self.m_staticTextXdegree = wx.StaticText( self.m_panel1, wx.ID_ANY, u"X����(��)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticShear.Add( self.m_staticTextXdegree, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlXdegree = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"30", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 90, 30 )
		sbSizerStaticShear.Add( self.m_spinCtrlXdegree, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticShear.Add((10, -1))
		
		self.m_staticTextYdegree = wx.StaticText( self.m_panel1, wx.ID_ANY, u"Y����(��)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticShear.Add( self.m_staticTextYdegree, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlYdegree = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"0", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 90, 0 )
		sbSizerStaticShear.Add( self.m_spinCtrlYdegree, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticShear.Add((10, -1))
		
		self.m_staticShearBkColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticShear.Add( self.m_staticShearBkColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )

		self.m_ShearColorPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.RED, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		sbSizerStaticShear.Add(self.m_ShearColorPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5)
				
		self.ctrlShearBkColor = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '(255, 0, 0, 255)', wx.DefaultPosition, wx.Size( 120,-1 ), wx.TE_PROCESS_ENTER )
		sbSizerStaticShear.Add( self.ctrlShearBkColor, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		self.ctrlShearBkColor.SetToolTipString(COLOR_FORMAT_TIP)
		self.ctrlShearBkColor.Bind(wx.EVT_TEXT, self.OnShearTxtChanged)
		bSizerSetting.Add( sbSizerStaticShear, 1, wx.EXPAND, 0 )
		
		#--------------- Mosaic starts ------------------------
		self.staticMosaic = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"����������" )
		sbSizerStaticMosaic = wx.StaticBoxSizer( self.staticMosaic, wx.HORIZONTAL )
		
		self.m_staticTextMosaicSize = wx.StaticText( self.m_panel1, wx.ID_ANY, u"�����˴�С(����)", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticMosaic.Add( self.m_staticTextMosaicSize, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_spinCtrlMosaicSize = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"8", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 2, 200, 8 )
		sbSizerStaticMosaic.Add( self.m_spinCtrlMosaicSize, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticMosaic.Add((10, -1))
		self.m_staticTextMosaicTip = wx.StaticText( self.m_panel1, label = u"�������ѡ��������Ĭ��������ͼƬ��" )
		self.m_staticTextMosaicTip.Enable( False )
		sbSizerStaticMosaic.Add( self.m_staticTextMosaicTip, 0, wx.ALIGN_CENTER|wx.ALL, 5 ) #wx.LEFT | wx.TOP, 22 )	
		
		bSizerSetting.Add( sbSizerStaticMosaic, 1, wx.EXPAND, 0 )
		
		#---------------- Mask starts -----------------------------
		self.staticMask = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�ɰ�Ч������" )
		sbSizerStaticMask = wx.StaticBoxSizer( self.staticMask, wx.HORIZONTAL )
		
		self.m_staticTextMask = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ѡ���ɰ�", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticMask.Add( self.m_staticTextMask, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlMaskPath = wx.TextCtrl( self.m_panel1, wx.ID_ANY, self.parent.mask_path, wx.DefaultPosition, (180, -1), wx.TE_READONLY)
		sbSizerStaticMask.Add( self.m_textCtrlMaskPath, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		self.mask = None         
		
		self.m_btnMaskBrowse = wx.Button( self.m_panel1, wx.ID_ANY, u"...", wx.DefaultPosition, (-1, 21), wx.BU_EXACTFIT )
		sbSizerStaticMask.Add( self.m_btnMaskBrowse, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		self.m_btnMaskBrowse.SetToolTipString('Ҫ��ָ�����ɰ��뵱ǰͼƬ��С��ͬ��\r\n�������������ʧ�档')
		
		self.m_staticTextMaskBg = wx.StaticText( self.m_panel1, wx.ID_ANY, u"������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticMask.Add( self.m_staticTextMaskBg, 0, wx.ALIGN_CENTER|wx.LEFT, 5 )
		
		self.m_MaskBgColorPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.WHITE, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		sbSizerStaticMask.Add(self.m_MaskBgColorPicker, 0, wx.ALIGN_CENTER|wx.LEFT, 5)
				
		self.ctrlMaskBgColor = wx.TextCtrl( self.m_panel1, wx.ID_ANY, '(255, 255, 255, 255)', wx.DefaultPosition, wx.Size( 120,-1 ), wx.TE_PROCESS_ENTER )
		sbSizerStaticMask.Add( self.ctrlMaskBgColor, 0, wx.ALIGN_CENTER|wx.LEFT | wx.RIGHT, 5 )
		
		self.ctrlMaskBgColor.SetToolTipString(COLOR_FORMAT_TIP)
		self.ctrlMaskBgColor.Bind(wx.EVT_TEXT, self.OnMaskBgTxtChanged)
		
		bSizerSetting.Add( sbSizerStaticMask, 1, wx.EXPAND, 0 )
		
		#---------------- Blend starts -----------------------------
		self.staticBlend = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"���Ч������" )
		sbSizerStaticBlend = wx.StaticBoxSizer( self.staticBlend, wx.HORIZONTAL )
		
		self.m_staticTextBlend = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ѡ��ͼƬ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticBlend.Add( self.m_staticTextBlend, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_textCtrlImg2Path = wx.TextCtrl( self.m_panel1, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, (220, -1), wx.TE_READONLY)
		sbSizerStaticBlend.Add( self.m_textCtrlImg2Path, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		self.img2 = None         
		
		self.m_btnBlendBrowse = wx.Button( self.m_panel1, wx.ID_ANY, u"...", wx.DefaultPosition, (-1, 21), wx.BU_EXACTFIT )
		sbSizerStaticBlend.Add( self.m_btnBlendBrowse, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		self.m_btnBlendBrowse.SetToolTipString('Ҫ��ָ����ͼƬ�뵱ǰͼƬ��С��ͬ��\r\n�������������ʧ�档')
		
		sbSizerStaticBlend.AddSpacer( ( 10, 0) )
		
		self.m_staticTextAlpha = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��͸����", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticBlend.Add( self.m_staticTextAlpha, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		self.m_spinCtrlAlpha = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"50", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 0, 100, 0 )
		sbSizerStaticBlend.Add( self.m_spinCtrlAlpha, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )		
		
		self.m_staticTextPercent = wx.StaticText( self.m_panel1, wx.ID_ANY, u"%", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticBlend.Add( self.m_staticTextPercent, 0, wx.ALIGN_CENTER|wx.RIGHT, 5 )
		
		bSizerSetting.Add( sbSizerStaticBlend, 1, wx.EXPAND, 0 )
		#------------------ split starts ---------------------------
		self.staticSplit = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"ͼƬ�ָ�" )
		sbSizerStaticSplit = wx.StaticBoxSizer( self.staticSplit, wx.HORIZONTAL )
		#sbSizerStaticEnhance.Add(self.staticEnhance, 0, wx.ALL|wx.RIGHT, 5 )
		
		#self.m_staticTextRow = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ˮƽ��Ŀ", wx.DefaultPosition, wx.DefaultSize, 0 )
		#self.m_staticTextRow.Wrap( -1 )
		#sbSizerStaticSplit.Add( self.m_staticTextRow, 0, wx.ALIGN_CENTER | wx.RIGHT, 5 )
		
		# add pixel option
		m_choicePixelChoicesX = [ u"ˮƽ��Ŀ", u"ˮƽ����" ]
		self.m_choicePixelX = wx.Choice( self.m_panel1, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, m_choicePixelChoicesX, 0 )
		self.m_choicePixelX.SetSelection( 0 )
		sbSizerStaticSplit.Add( self.m_choicePixelX, 0, wx.ALL | wx.ALIGN_CENTER, 5 )
		
		self.m_spinCtrlRow = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"2", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 1, 10000, 2 )
		sbSizerStaticSplit.Add( self.m_spinCtrlRow, 0, wx.ALL | wx.ALIGN_CENTER, 5 )
		
		sbSizerStaticSplit.Add( ( 15, -1) )
		
		#self.m_staticTextCol = wx.StaticText( self.m_panel1, wx.ID_ANY, u"��ֱ��Ŀ", wx.DefaultPosition, wx.DefaultSize, 0 )
		#self.m_staticTextCol.Wrap( -1 )
		#sbSizerStaticSplit.Add( self.m_staticTextCol, 0, wx.ALIGN_CENTER | wx.RIGHT, 5 )
		m_choicePixelChoicesY = [ u"��ֱ��Ŀ", u"��ֱ����" ]
		self.m_choicePixelY = wx.Choice( self.m_panel1, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, m_choicePixelChoicesY, 0 )
		self.m_choicePixelY.SetSelection( 0 )
		sbSizerStaticSplit.Add( self.m_choicePixelY, 0, wx.ALL | wx.ALIGN_CENTER, 5 )
		
		self.m_spinCtrlCol = wx.SpinCtrl( self.m_panel1, wx.ID_ANY, u"2", wx.DefaultPosition, wx.Size( 60,-1 ), wx.SP_ARROW_KEYS, 1, 10000, 2 )
		sbSizerStaticSplit.Add( self.m_spinCtrlCol, 0, wx.ALL | wx.ALIGN_CENTER, 5 )	
		
		sbSizerStaticSplit.Add((5, -1))
		
		self.m_staticLineColor = wx.StaticText( self.m_panel1, wx.ID_ANY, u"Ԥ��������ɫ", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticSplit.Add( self.m_staticLineColor, 0, wx.ALL | wx.ALIGN_CENTER|wx.LEFT, 5 )

		self.m_LineColourPicker = wx.ColourPickerCtrl( self.m_panel1, wx.ID_ANY, wx.WHITE, wx.DefaultPosition, wx.Size( 50,12 ), wx.CLRP_DEFAULT_STYLE)
		sbSizerStaticSplit.Add(self.m_LineColourPicker, 0, wx.ALL | wx.ALIGN_CENTER|wx.LEFT, 5)				
	
		self.m_LineColourPicker.SetToolTipString(SPLIT_LINE_COLOR_TIP)

		bSizerSetting.Add( sbSizerStaticSplit, 1, wx.EXPAND, 0 )
		# ------------------------------------ slice gone ----------------------				
		self.staticSliceGone = wx.StaticBox( self.m_panel1, wx.ID_ANY, u"�м�ȥ��" )
		sbSizerStaticSliceGone = wx.StaticBoxSizer( self.staticSliceGone, wx.HORIZONTAL )
		
		self.sliceModeStyleList = [u'����', u'����']
		self.m_checkBoxSliceStyle = wx.ComboBox(self.m_panel1, value = self.sliceModeStyleList[0], choices = self.sliceModeStyleList, style = wx.CB_READONLY )
		sbSizerStaticSliceGone.Add( self.m_checkBoxSliceStyle, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		self.m_staticSlice = wx.StaticText( self.m_panel1, wx.ID_ANY, u"ƴ��", wx.DefaultPosition, wx.DefaultSize, 0 )
		sbSizerStaticSliceGone.Add( self.m_staticSlice, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		
		sbSizerStaticSliceGone.Add((10, -1))
		
		self.m_staticSliceModeTip = wx.StaticText( self.m_panel1, label = u"��ʹ����꣬ѡ��Ҫȥ�����м䲿��" )
		self.m_staticSliceModeTip.Enable( False )
		sbSizerStaticSliceGone.Add( self.m_staticSliceModeTip, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
		bSizerSetting.Add( sbSizerStaticSliceGone, 1, wx.EXPAND, 0 )

		#--------------the bottom buttons starts -------------------
		bSizerMain.Add( bSizerSetting, 1, wx.EXPAND, 5 )
		
		bSizerBtnBottom = wx.BoxSizer( wx.HORIZONTAL )  
		bSizerBtnBottom.AddSpacer( ( 0, 0), 1, wx.EXPAND, 5 )
		
		self.m_preview = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.preview_ico, wx.DefaultPosition, wx.Size(48, 48), wx.NO_BORDER |wx.BU_EXACTFIT  )
		bSizerBtnBottom.Add( self.m_preview, 0, wx.ALL|wx.RIGHT, 5 )
		self.m_preview.SetToolTipString(u'Ԥ��Ч��')
		
		self.m_reset = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.refresh_ico, wx.DefaultPosition, wx.Size(48, 48), wx.NO_BORDER|wx.BU_EXACTFIT  )
		bSizerBtnBottom.Add( self.m_reset, 0, wx.ALL|wx.RIGHT, 5 )
		self.m_reset.SetToolTipString(u'���¼���')
		
		self.m_close = wx.BitmapButton( self.m_panel1, wx.ID_ANY, self.parent.close32_ico, wx.DefaultPosition, wx.Size(48, 48), wx.NO_BORDER|wx.BU_EXACTFIT  )
		bSizerBtnBottom.Add( self.m_close, 0, wx.ALL|wx.RIGHT, 5 )
		self.m_close.SetToolTipString(u'�ر�')
		
		bSizerMain.Add( bSizerBtnBottom, 1, wx.EXPAND, 5 )
		
		self.lb_info = wx.StaticText(self, label=u'������ʾ����ЩЧ����Ҫ�ϳ�ʱ��, �����ĵȴ���')
		self.lb_info.Enable(False)
		bSizerMain.Add(self.lb_info, 0, wx.ALL|wx.CENTER, 5 )
		#--------------the bottom buttons starts -------------------		
		
		self.m_panel1.SetSizer( bSizerMain )
		self.m_panel1.Layout()
		bSizerMain.Fit( self.m_panel1 )
		bSizer1.Add( self.m_panel1, 1, wx.EXPAND |wx.ALL, 5 )

		self.SetSizer( bSizer1 )                
		self.Layout()           
		self.Centre( wx.BOTH )
		self.RefreshTitle()
		
		# Connect Events
		self.m_buttonDel.Bind( wx.EVT_BUTTON, self.OnDel)
		self.m_buttonDelAll.Bind( wx.EVT_BUTTON, self.OnDelAll)
		self.m_btnUP.Bind( wx.EVT_BUTTON, self.OnBtnUp )
		self.m_btnDown.Bind( wx.EVT_BUTTON, self.OnBtnDown )
		self.m_buttonRound.Bind( wx.EVT_TOGGLEBUTTON, self.OnRoundClick )
		self.m_buttonBackground.Bind( wx.EVT_TOGGLEBUTTON, self.OnBackgroundClick )
		self.m_bpButtonPlus.Bind( wx.EVT_BUTTON, self.OnPlus )
		self.m_checkBoxDefault.Bind( wx.EVT_CHECKBOX, self.OnCheckDefault )
		#self.m_checkBoxIsCut.Bind (wx.EVT_CHECKBOX, self.OnCheckIsCut )
		self.m_radioBtnCkColor.Bind( wx.EVT_RADIOBUTTON, self.OnCheckColor )
		self.m_radioBtnCkPic.Bind( wx.EVT_RADIOBUTTON, self.OnCheckPic )
		self.m_btnBrowse.Bind( wx.EVT_BUTTON, self.OnBtnBrowse )
		self.m_btnMaskBrowse.Bind (wx.EVT_BUTTON, self.OnBtnMaskBrowse )
		self.m_btnBlendBrowse.Bind ( wx.EVT_BUTTON, self.OnBtnBlendBrowse )
		self.m_preview.Bind( wx.EVT_BUTTON, self.OnPreview )
		self.m_reset.Bind( wx.EVT_BUTTON, self.OnReset )
		self.m_close.Bind( wx.EVT_BUTTON, self.OnClose )
		#self.m_bgColourPicker.Bind(wx.EVT_COLOURPICKER_CHANGED, self.OnBgColorChanged)
		self.m_buttonBorder.Bind(wx.EVT_TOGGLEBUTTON, self.OnBorderClick)
		#self.m_borderColourPicker.Bind(wx.EVT_COLOURPICKER_CHANGED, self.OnBorderColorChanged)
		self.m_buttonShadow.Bind(wx.EVT_TOGGLEBUTTON, self.OnShadowClick)
		#self.m_ReflectionBgColourPicker.Bind(wx.EVT_COLOURPICKER_CHANGED, self.OnReflectionColorChanged)
		self.m_buttonReflection.Bind(wx.EVT_TOGGLEBUTTON, self.OnReflectionClick)
		self.m_buttonEnhance.Bind(wx.EVT_TOGGLEBUTTON, self.OnEnhanceClick)
		self.m_buttonFilter.Bind(wx.EVT_TOGGLEBUTTON, self.OnFilterClick)
		self.m_buttonRotate.Bind(wx.EVT_TOGGLEBUTTON, self.OnRotateClick)
		self.m_buttonMirror.Bind(wx.EVT_TOGGLEBUTTON, self.OnMirrorClick)
		self.m_buttonInvert.Bind(wx.EVT_TOGGLEBUTTON, self.OnInvertClick)
		#self.m_buttonMolten.Bind(wx.EVT_TOGGLEBUTTON, self.OnMoltenClick)
		self.m_buttonLighting.Bind(wx.EVT_TOGGLEBUTTON, self.OnLightingClick)
		self.m_buttonSketch.Bind(wx.EVT_TOGGLEBUTTON, self.OnSketchClick)
		#self.m_buttonIce.Bind(wx.EVT_TOGGLEBUTTON, self.OnIceClick)
		self.m_buttonPapercut.Bind(wx.EVT_TOGGLEBUTTON, self.OnPapercutClick)
		#self.m_buttonTan.Bind(wx.EVT_TOGGLEBUTTON, self.OnTanClick)
		self.m_buttonSwirl.Bind(wx.EVT_TOGGLEBUTTON, self.OnSwirlClick)
		#self.m_buttonSpherize.Bind(wx.EVT_TOGGLEBUTTON, self.OnSpherizeClick)
		self.m_buttonDeYellow.Bind(wx.EVT_TOGGLEBUTTON, self.OnDeYellowClick)				
		self.m_radioBtnAny.Bind(wx.EVT_RADIOBUTTON, self.OnBtnAnyClick)
		self.m_radioBtnFixed.Bind(wx.EVT_RADIOBUTTON, self.OnBtnFixedClick)
		self.m_buttonPuzzle.Bind(wx.EVT_TOGGLEBUTTON, self.OnPuzzleClick)
		self.m_buttonFocus.Bind(wx.EVT_TOGGLEBUTTON, self.OnFocusClick)
		self.m_buttonCornerFold.Bind(wx.EVT_TOGGLEBUTTON, self.OnCornerFoldClick)
		self.m_buttonTear.Bind(wx.EVT_TOGGLEBUTTON, self.OnTearClick)
		self.m_buttonFade.Bind(wx.EVT_TOGGLEBUTTON, self.OnFadeClick)
		self.m_buttonMix.Bind(wx.EVT_TOGGLEBUTTON, self.OnMixClick)
		self.m_buttonReplace.Bind(wx.EVT_TOGGLEBUTTON, self.OnReplaceClick)
		self.m_MixColourPicker.Bind (wx.EVT_COLOURPICKER_CHANGED, self.OnMixColorChanged)
		self.m_OldColourPicker.Bind (wx.EVT_COLOURPICKER_CHANGED, self.OnOldColorChanged)
		self.m_OldColourMaxPicker.Bind (wx.EVT_COLOURPICKER_CHANGED, self.OnOldColorMaxChanged)
		self.m_NewColourPicker.Bind (wx.EVT_COLOURPICKER_CHANGED, self.OnNewColorChanged)
		self.m_MaskBgColorPicker.Bind (wx.EVT_COLOURPICKER_CHANGED, self.OnMaskBgColorChanged)
		self.m_buttonAutoBrightness.Bind(wx.EVT_TOGGLEBUTTON, self.OnAutoBrightnessClick)
		self.m_ShearColorPicker.Bind (wx.EVT_COLOURPICKER_CHANGED, self.OnShearColorChanged)
		self.m_buttonShear.Bind(wx.EVT_TOGGLEBUTTON, self.OnShearClick)
		self.m_buttonRoll.Bind(wx.EVT_TOGGLEBUTTON, self.OnRollClick)
		self.m_buttonMosaic.Bind(wx.EVT_TOGGLEBUTTON, self.OnMosaicClick)
		self.m_buttonMask.Bind(wx.EVT_TOGGLEBUTTON, self.OnMaskClick)
		self.m_buttonBlend.Bind(wx.EVT_TOGGLEBUTTON, self.OnBlendClick)
		self.m_buttonWaterMark.Bind(wx.EVT_TOGGLEBUTTON, self.OnWaterMarkClick)
		self.m_buttonSplit.Bind(wx.EVT_TOGGLEBUTTON, self.OnSplitClick)
		self.m_buttonSlice.Bind(wx.EVT_TOGGLEBUTTON, self.OnSliceClick)
		self.m_buttonListView.Bind( wx.EVT_BUTTON, self.OnListView )
		
		#self.Bind (wx.EVT_LEAVE_WINDOW, self.OnLeave)
		
		# Toggle the following controls
		self.toggleCheckBox(self.m_checkBoxDefault)
		self.toggleRoundPanel(False)
		self.toggleBackgroundPanel(False)
		self.toggleRadioBox()
		self.toggleBorderPanel(False)
		self.toggleShadowPanel(False)
		self.toggleReflectionPanel(False)
		self.toggleEnhancePanel(False)
		self.toggleFilterPanel(False)		
		self.toggleRotatePanel(False)
		self.toggleMirrorPanel(False)
		self.toggleInvertPanel(False)
		self.toggleLightingPanel(False)
		self.toggleSketchPanel(False)
		self.togglePapercutPanel(False)
		self.toggleSwirlPanel(False)		
		#self.toggleSpherizePanel(False)
		self.toggleDeYellowPanel(False)
		self.togglePuzzlePanel(False)
		self.toggleFocusPanel(False)
		self.toggleCornerFoldPanel(False)
		self.toggleTearPanel(False)
		self.toggleFadePanel(False)
		self.toggleMixPanel (False)
		self.toggleReplacePanel (False)
		self.toggleAutoBrightnessPanel (False )
		self.toggleShearPanel (False)
		self.toggleRollPanel (False)
		self.toggleMosaicPanel(False)
		self.toggleMaskPanel(False)
		self.toggleBlendPanel (False)
		self.toggleSplitPanel (False)
		self.toggleSlicePanel (False)
		
		ctrl_focus_list = [self.ctrlOldColor, self.ctrlOldColorMax, self.ctrlNewColor, self.ctrlShearBkColor, self.ctrlMaskBgColor, self.ctrlColor ]
		for ctrl in ctrl_focus_list: ctrl.Bind(wx.EVT_LEFT_DOWN, self.OnCtrlClick)
	
	def OnListView( self, event ):
		if not self.list_view_dlg:
			self.list_view_dlg = ShowListDlg (self)
		self.list_view_dlg.addlines(self.parent.picPaths)
		self.list_view_dlg.ShowModal()
		
		event.Skip()
	
	def  OnCtrlClick (self, event):
		ctrl = event.GetEventObject()
		if self.ctrlOnFocus: self.ctrlOnFocus.SetSelection(0, 0)
		self.ctrlOnFocus = ctrl
		event.Skip()
	
	def SetPickedColor (self, X, Y): #color):
		"""
		Set value passed from Main UI
		"""
		#if self.ctrlOnFocus: self.ctrlOnFocus.SetValue(color)
		if not self.parent.imageLabel.GetLabel().startswith(u"(Ԥ��)"):
			img = self.parent.imageOpen(self.parent.getCurJpg()).convert("RGBA")
		else:
			img = self.parent.im_preview
		pixels = img.load()
		if self.ctrlOnFocus:  self.ctrlOnFocus.SetValue(str(pixels[X, Y]))

	def RefreshTitle (self):
		self.SetTitle(self.parent.getGuiTitle(u"ͼƬ��Ч"))
		im = self.parent.imageOpen(self.parent.picPaths[self.parent.currentPicture]).convert("RGBA")
		brightness_current = str(int(brightness(im)))
		self.ctrlCurBrightness.SetValue(brightness_current)
		self.m_spinCtrlMosaicSize.SetRange(2, min(im.size[0], im.size[1]))
		self.m_spinCtrlMosaicSize.SetValue(min(8, (min(im.size[0], im.size[1]))))
	
	def OnDel( self, event ):
		sel = self.m_checkList1.GetSelection()
		if sel != -1:
			self.m_checkList1.Delete(sel)
			if len(self.m_checkList1.Items) > 0:
				self.m_checkList1.SetSelection(len(self.m_checkList1.Items) - 1 )
		event.Skip()
		
	def OnDelAll( self, event ):
		self.m_checkList1.Clear()
		event.Skip()
	
	# Virtual event handlers, overide them in your derived class
	def OnBtnUp( self, event ):
		sel = self.m_checkList1.GetSelection()
		items = self.m_checkList1.Items
		if sel > 0: # selected not -1 And not zero
			self.moveUpDown(1)              
		event.Skip()
	def moveUpDown(self, row):
		sel = self.m_checkList1.GetSelection()
		sel_string = self.m_checkList1.GetString(sel)
		items = self.m_checkList1.Items         
		if sel != len(items) + row:
			self.m_checkList1.Delete(sel)
			self.m_checkList1.Insert(sel_string, sel - row)
			self.m_checkList1.SetSelection(sel - row)
	
	def OnBtnDown( self, event ):
		sel = self.m_checkList1.GetSelection()
		items = self.m_checkList1.Items
		if sel not in [ -1, len(items) - 1]:
			self.moveUpDown(-1)
		event.Skip()
	def noOpenSetting(self, isPressed):
		if isPressed and  self.settingCounts == 2:
			wx.MessageBox("��Ǹ���ﵽ������������򿪵��������", "С����", wx.OK | wx.ICON_INFORMATION)                  
			return True
		else:
			return False
	def OnSettingClick( self, event, togglePanel, tooltip ):                
		btn = event.GetEventObject()
		isPressed = btn.GetValue()
		if not self.noOpenSetting(isPressed):
			togglePanel(isPressed)
			if isPressed:
				btn.UnsetToolTip()
				self.settingCounts += 1
			else:
				btn.SetToolTipString(tooltip)
				self.settingCounts -= 1
		else:
			btn.SetValue(False)
		event.Skip()
		
	def OnRoundClick( self, event ):
		self.OnSettingClick(event, self.toggleRoundPanel, u'��Բ������')
		event.Skip()
		
	def toggleRoundPanel(self, isPressed):
		# Բ��Ч����
		ctrl_list = [self.m_checkBoxDefault, self.m_staticTextRound, self.m_spinCtrlNum, \
		self.m_staticTextSize, self.staticRound ]
		self.toggleSettingPanel (ctrl_list, isPressed)

	def toggleSettingPanel(self, ctrl_list, isPressed):		
		for ctrl in ctrl_list: ctrl.Show(isPressed)
		self.Layout()
	
	def OnBackgroundClick( self, event ):
		self.OnSettingClick(event, self.toggleBackgroundPanel, BACKGROUND_TIP)
		event.Skip()
		
	def toggleBackgroundPanel (self, isPressed):
		# ������           
		ctrl_list = [self.m_radioBtnCkColor, self.m_bgColourPicker, self.m_radioBtnCkPic, self.m_textCtrlPath, \
		self.m_btnBrowse, self.m_staticTextHtxt, self.m_textCtrlHbox, self.m_staticTextVtxt, self.m_textCtrlVbox, \
		self.staticBackground, self.m_comboTextHPos, self.m_comboTextVPos, self.m_staticTextHPos, self.m_staticTextVPos, \
		self.m_methodText, self.m_comboMethod ]
		self.toggleSettingPanel (ctrl_list, isPressed)
	def OnMosaicClick( self, event ):
		self.OnSettingClick(event, self.toggleMosaicPanel, u'������������')
		event.Skip()
		
	def	toggleMosaicPanel(self, isPressed):
		#Mosaic
		ctrl_list = [self.staticMosaic, self.m_staticTextMosaicSize, self.m_spinCtrlMosaicSize, self.m_staticTextMosaicTip ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnMaskClick( self, event ):
		self.OnSettingClick(event, self.toggleMaskPanel, u'���ɰ�����')
		event.Skip()
		
	def OnWaterMarkClick (self, event ):
		#self.OnSettingClick(event, self.toggleTxtDlg, u'��ˮӡ����')
		btn = event.GetEventObject()
		isPressed = btn.GetValue()

		if isPressed: btn.UnsetToolTip()
		else:	btn.SetToolTipString(u'��ˮӡ����')

		self.toggleTxtDlg(isPressed)
		event.Skip()
		
	def toggleTxtDlg (self, isPressed):
		if isPressed:
			if self.parent.txtDlg: self.parent.txtDlg.Show()
			else: self.doTxtDlg()
	

	def toggleMaskPanel(self, isPressed):
		#Mosaic
		ctrl_list = [self.staticMask, self.m_staticTextMask, self.m_textCtrlMaskPath, self.m_btnMaskBrowse, self.m_staticTextMaskBg, \
					 self.m_MaskBgColorPicker, self.ctrlMaskBgColor ]
		self.toggleSettingPanel (ctrl_list, isPressed)		
		
	def OnBlendClick (self, event ):
		self.OnSettingClick(event, self.toggleBlendPanel, u'�򿪻������')
		event.Skip()

	def OnSplitClick (self, event ):
		self.OnSettingClick(event, self.toggleSplitPanel, SPLIT_TIP)
		event.Skip()
		
	def OnSliceClick (self, event ):
		self.OnSettingClick(event, self.toggleSlicePanel, SLICE_TIP)
		event.Skip()
		
	def toggleSplitPanel(self, isPressed):
		#Mosaic
		ctrl_list = [self.staticSplit, self.m_choicePixelX , self.m_spinCtrlRow, self.m_choicePixelY, self.m_spinCtrlCol, \
					 self.m_staticLineColor, self.m_LineColourPicker ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def toggleSlicePanel(self, isPressed):
		#Mosaic
		ctrl_list = [self.staticSliceGone, self.m_checkBoxSliceStyle , self.m_staticSliceModeTip, self.m_staticSlice ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def toggleBlendPanel(self, isPressed):
		#Mosaic
		ctrl_list = [self.staticBlend, self.m_staticTextBlend, self.m_textCtrlImg2Path, self.m_btnBlendBrowse, self.m_staticTextAlpha, self.m_spinCtrlAlpha, \
					 self.m_staticTextPercent ]
		self.toggleSettingPanel (ctrl_list, isPressed)	
		
	def OnBorderClick(self, event):
		self.OnSettingClick(event, self.toggleBorderPanel, u'�򿪱߿�����')
		event.Skip()
		
	def toggleBorderPanel (self, isPressed):
		# �߿�
		ctrl_list = [self.staticBorder, self.m_staticTextBorder, self.m_textCtrlBorderBox, self.m_staticTextColor, \
		self.m_borderColourPicker, self.m_staticTextOpacity, self.m_sliderOpacity ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnShadowClick(self, event):
		self.OnSettingClick(event, self.toggleShadowPanel, u'����Ӱ����')
		event.Skip()

	def toggleShadowPanel(self, isPressed):
		#��Ӱ
		ctrl_list = [self.m_staticHTextShadow, self.m_textCtrlShadowHBox, self.m_staticVTextShadow, self.m_textCtrlShadowVBox, self.staticShadow ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnRollClick(self, event):
		self.OnSettingClick(event, self.toggleRollPanel, u'�򿪹�������')
		event.Skip()

	def toggleRollPanel(self, isPressed):
		#����
		ctrl_list = [self.m_staticHTextRoll, self.m_textCtrlRollHBox, self.m_staticVTextRoll, self.m_textCtrlRollVBox, self.staticRoll ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def toggleShearPanel (self, isPressed):
		# ��б
		ctrl_list = [self.staticShear, self.m_staticTextXdegree, self.m_spinCtrlXdegree, self.m_staticTextYdegree, \
			     self.m_spinCtrlYdegree, self.m_staticShearBkColor, self.m_ShearColorPicker, self.ctrlShearBkColor ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnShearClick(self, event):
		self.OnSettingClick(event, self.toggleShearPanel, u'����б����')
		event.Skip()
		
	def OnReflectionClick(self, event):
		self.OnSettingClick(event, self.toggleReflectionPanel, u'�򿪵�Ӱ����')
		event.Skip()
	
	def toggleReflectionPanel (self, isPressed):
		ctrl_list = [self.staticReflection, self.m_staticTextDepth, self.m_textCtrlDepthBox, self.m_staticReflectionBgColor, \
		self.m_ReflectionBgColourPicker, self.m_staticReflectionOpacity, self.m_sliderReflectionOpacity ]
		self.toggleSettingPanel (ctrl_list, isPressed)	
		
	def OnEnhanceClick( self, event ):
		#ͼ����ǿ
		self.OnSettingClick(event, self.toggleEnhancePanel, u'��ͼ����ǿ����')
		event.Skip()
		
	def toggleEnhancePanel (self, isPressed):
		# ��ǿ��
		ctrl_list = [self.staticEnhance, self.m_staticTextFactor, self.m_spinCtrlFactor, self.m_staticTextEnhance, self.m_EnhanceBox ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnFilterClick( self, event ):
		self.OnSettingClick(event, self.toggleFilterPanel, u'���˾�����')
		event.Skip()
		
	def toggleFilterPanel (self, isPressed):
		# �˾�
		for checkbox in self.m_FilterBoxList:
			checkbox.Show(isPressed)        
		self.staticFilter.Show(isPressed)
		self.Layout()
		
	def OnRotateClick( self, event ):
		self.OnSettingClick(event, self.toggleRotatePanel, u'����ת����')
		event.Skip()
		
	def toggleRotatePanel (self, isPressed):
		# ��ת
		ctrl_list = [self.staticRotate, self.m_radioBtnFixed, self.m_choiceFixed, self.m_radioBtnAny, self.m_spinCtrlAny ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnMirrorClick( self, event ):
		self.OnSettingClick(event, self.toggleMirrorPanel, u'�򿪾���ת����')
		event.Skip()
		
	def toggleMirrorPanel (self, isPressed):
		# ����ת  
		self.staticMirror.Show(isPressed)
		self.m_staticTextMirror.Show(isPressed)
		self.m_MirrorBox.Show(isPressed)
		self.Layout()		
		
	def OnInvertClick( self, event ):
		self.OnSettingClick(event, self.toggleInvertPanel, u'�򿪵�Ƭ����')
		event.Skip()
		
	def toggleInvertPanel (self, isPressed):
		# ����ת  
		self.staticInvert.Show(isPressed)
		self.m_staticTextInvert.Show(isPressed)
		self.m_spinCtrlInvert.Show(isPressed)
		self.Layout()
		
	def OnLightingClick( self, event ):
		self.OnSettingClick(event, self.toggleLightingPanel, u'�򿪹�������')
		event.Skip()
		
	def toggleLightingPanel (self, isPressed):
		# ����
		ctrl_list = [self.staticLighting, self.m_staticTextLightingPos, self.m_comboTextLightingPos, self.m_staticTextLighting, self.m_spinCtrlLighting ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnSketchClick( self, event ):
		self.OnSettingClick(event, self.toggleSketchPanel, u'����������')
		event.Skip()
		
	def toggleSketchPanel (self, isPressed):
		# ����  
		self.staticSketch.Show(isPressed)
		self.m_staticTextSketch.Show(isPressed)
		self.m_spinCtrlSketch.Show(isPressed)
		self.Layout()
		
	def OnPapercutClick( self, event ):
		self.OnSettingClick(event, self.togglePapercutPanel, u'�򿪼�ֽ����')
		event.Skip()
		
	def togglePapercutPanel (self, isPressed):
		# ��ֽ
		ctrl_list = [self.staticPapercut, self.m_staticPapercutBgColor, self.m_PapercutBgColourPicker, self.m_staticPapercutFgColor, \
		self.m_PapercutFgColourPicker, self.m_staticTextPapercut, self.m_spinCtrlPapercut ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnPuzzleClick( self, event ):
		self.OnSettingClick(event, self.togglePuzzlePanel, PUZZLE_TIP)
		event.Skip()
		
	def togglePuzzlePanel (self, isPressed):
		# ƴͼ
		ctrl_list = [self.staticPuzzle, self.m_staticTextQtyX, self.m_staticTextQtyY, self.m_spinCtrlQtyX, self.m_spinCtrlQtyY, self.m_buttonListView, \
		self.m_PuzzleSrcBox, self.m_staticPuzzleBgColor, self.m_PuzzleBgColourPicker, self.m_PuzzleOrderBox, self.m_spinCtrlGap, self.m_staticTextGap, self.m_checkBoxResize ]		
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnTearClick( self, event ):               
		self.OnSettingClick(event, self.toggleTearPanel, TEAR_TIP)
		event.Skip()
		
	def toggleTearPanel (self, isPressed):
		# ��ֽ
		ctrl_list = [self.staticTear, self.m_staticTearDirection, self.m_TearPosBox, self.m_staticTextTear, self.m_spinCtrlTear, self.m_checkBoxIsCut ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnMixClick( self, event ):               
		self.OnSettingClick(event, self.toggleMixPanel, u'�򿪻�ɫ����')
		event.Skip()
		
	def OnReplaceClick( self, event ):
		self.OnSettingClick(event, self.toggleReplacePanel, REPLACE_TIP)
		event.Skip()
		
	def toggleReplacePanel (self, isPressed):
		# ��ɫ�滻
		ctrl_list = [self.staticReplace, self.m_checkBoxReplaceStyle, self.m_staticTextModeTip, self.m_staticOldColor, self.m_ColorCompareBox, self.m_OldColourPicker, self.ctrlOldColor, \
		self.m_staticNewColor, self.m_NewColourPicker, self.ctrlNewColor, self.ctrlOldColor ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		ctrl_list_range = [ self.m_staticTextOldMax, self.m_OldColourMaxPicker, self.ctrlOldColorMax]

		flag = self.m_ColorCompareBox.GetValue()
		isRange = flag == u'��'
		self.toggleSettingPanel (ctrl_list_range, isPressed and isRange)
		
	def OnColorComboChanged (self, event):
		self.toggleReplacePanel (True)
	
	def OnReplaceModeChanged(self, event):
		isFill = (self.m_checkBoxReplaceStyle.Selection == 1)
		ctrl_list = [self.m_staticOldColor, self.m_OldColourPicker, self.m_ColorCompareBox, self.ctrlOldColor]
		[ oldCtrls.Enable(not isFill) for oldCtrls in ctrl_list ]
		ctrl_list_range = [ self.m_staticTextOldMax, self.m_OldColourMaxPicker, self.ctrlOldColorMax]
		
		flag = self.m_ColorCompareBox.GetValue()
		isRange = flag == u'��'
		self.toggleSettingPanel (ctrl_list_range, isRange)
		
	def toggleMixPanel (self, isPressed):
		# ��ɫ
		ctrl_list = [self.staticMix, self.m_staticTextMixFactor, self.m_spinCtrlMixFactor, \
		self.m_staticMixColor, self.m_MixColourPicker, self.ctrlColor ]
		self.toggleSettingPanel (ctrl_list, isPressed)	
		
	def OnFadeClick( self, event ):
		self.OnSettingClick(event, self.toggleFadePanel, u'�򿪽�������')
		event.Skip()
		
	def toggleFadePanel (self, isPressed):
		# ����
		ctrl_list = [self.staticFade, self.m_staticTextFadePos, self.m_spinCtrlFadePos, self.m_staticFadeMode, \
		self.m_FadeModeBox, self.m_staticFadeColor, self.m_FadeColourPicker ]
		self.toggleSettingPanel (ctrl_list, isPressed)
	
	def OnFocusClick( self, event ):
		# �۹�
		self.OnSettingClick(event, self.toggleFocusPanel, FOCUS_TIP )
		event.Skip()
	def OnAutoBrightnessClick (self, event):
		# �Զ�����
		self.OnSettingClick(event, self.toggleAutoBrightnessPanel, u'���Զ���������\r\n����ͼƬ��ָ��������' )
		event.Skip()
		
	def toggleFocusPanel (self, isPressed):
		# �۹�
		ctrl_list = [self.staticFocus, self.m_staticTextRadiusIn, self.m_spinCtrlRadiusIn, self.m_staticTextRadiusOut, \
		self.m_spinCtrlRadiusOut, self.m_staticFocusEdgeColor, self.m_FocusEdgeColourPicker, self.m_checkBoxFocusEdge ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def toggleAutoBrightnessPanel (self, isPressed):
		#�Զ�����
		ctrl_list = [self.staticAutoBrightness, self.m_staticCurBrightness, self.ctrlCurBrightness, \
					 self.m_staticNewBrightness, self.m_spinCtrlNewBrightness ]
		self.toggleSettingPanel (ctrl_list, isPressed)
		
	def OnSwirlClick( self, event ):
		self.OnSettingClick(event, self.toggleSwirlPanel, u'����������')
		event.Skip()
		
	def toggleSwirlPanel (self, isPressed):
		# ���� 
		self.staticSwirl.Show(isPressed)
		self.m_staticTextSwirl.Show(isPressed)
		self.m_sliderSwirlOpacity.Show(isPressed)
		self.Layout()
		
	def OnDeYellowClick( self, event ):
		self.OnSettingClick(event, self.toggleDeYellowPanel, u'������Ƭ����')
		event.Skip()
		
	def toggleDeYellowPanel (self, isPressed):
		# ����Ƭ 
		self.staticDeYellow.Show(isPressed)
		self.m_radioBoxDeYellow_1.Show(isPressed)
		self.m_radioBoxDeYellow_2.Show(isPressed)
		self.Layout()
		
	def OnPlus( self, event ):
		for btn in self.taskListBtns:
			if btn.GetValue():
				self.m_checkList1.Append(btn.Label)
		event.Skip()
	
	def OnBtnFixedClick(self, event):
		self.m_choiceFixed.Enable(True)
		self.m_spinCtrlAny.Enable(False)
		event.Skip()
		
	def OnBtnAnyClick(self, event):
		self.m_spinCtrlAny.Enable(True)
		self.m_choiceFixed.Enable(False)
		event.Skip()
		
	def toggleCheckBox(self, button):               
		self.m_spinCtrlNum.Enable( not button.GetValue() )
		if button.GetValue(): self.m_spinCtrlNum.SetValue(5)
		
	def toggleRadioBox(self):
		isColorChecked = self.m_radioBtnCkColor.GetValue()
		self.m_bgColourPicker.Enable(isColorChecked)            
		#self.m_textCtrlPath.Enable(not isColorChecked)
		self.m_btnBrowse.Enable(not isColorChecked)
		self.m_methodText.Enable(not isColorChecked)
		self.m_comboMethod.Enable(not isColorChecked)
		self.offsetGroupStatus()
		
	def OnCornerFoldClick( self, event ):
		self.OnSettingClick(event, self.toggleCornerFoldPanel, u'���۽�����\r\nע�⣺ֻ��PNG��GIF����͸��Ч��')
		event.Skip()
		
	def toggleCornerFoldPanel (self, isPressed):
		# ����Ƭ
		ctrl_list = [self.staticCornerFold, self.m_staticTextEdge, self.m_spinCtrlEdge, self.m_staticTextWhitePercent, \
		self.m_spinCtrlWhitePercent, self.m_staticFoldPos, self.m_FoldPosBox ]
		self.toggleSettingPanel (ctrl_list, isPressed)		

	def OnCheckDefault( self, event ):
		btn = event.GetEventObject()            
		self.toggleCheckBox(self.m_checkBoxDefault)             
		event.Skip()
		
	def OnCheckPic(self, event):
		self.toggleRadioBox()
		event.Skip()
		
	def OnCheckColor(self, event):
		self.toggleRadioBox()
		event.Skip()            
	
	def KeyDown(self, event):
		curSel = event.GetSelection()
		event.Skip()
		
	def OnBtnBrowse( self, event ):			
		fd = wx.FileDialog(self, u'��ѡ��һ�ű���ͼƬ', self.m_textCtrlPath.GetValue(), wx.EmptyString, wildcard, wx.OPEN)
		if fd.ShowModal() == wx.ID_OK:
			self.pic = fd.GetPath()
			self.m_textCtrlPath.SetValue(self.pic)
			self.m_textCtrlPath.SetToolTip(wx.ToolTip(self.pic))
		fd.Destroy()
		event.Skip()
		
	def OnBtnMaskBrowse( self, event ):
		fd = wx.FileDialog(self, u'��ѡ��һ���ɰ�ͼƬ', self.m_textCtrlPath.GetValue(), wx.EmptyString, wildcard, wx.OPEN)
		if fd.ShowModal() == wx.ID_OK:			
			try:
				path = fd.GetPath()
				self.mask = Image.open (path)
				self.m_textCtrlMaskPath.SetValue(path)
				self.m_textCtrlMaskPath.SetToolTipString (path)
				self.parent.mask_path = path	
			except:
				wx.MessageBox("��Ǹ�����ɰ�ʱ���������ԡ�", "С����", wx.OK | wx.ICON_INFORMATION) 
		fd.Destroy()
		event.Skip()
		
	def OnBtnBlendBrowse( self, event ):			
		fd = wx.FileDialog(self, u'��ѡ��һ��ͼƬ', wx.EmptyString, wx.EmptyString, wildcard, wx.OPEN)
		if fd.ShowModal() == wx.ID_OK:
			try:
				path = fd.GetPath()
				self.img2 = Image.open (path)
				self.m_textCtrlImg2Path.SetValue(path)
				self.m_textCtrlImg2Path.SetToolTipString(path)
			except:
				wx.MessageBox("��Ǹ����ͼƬʱ���������ԡ�", "С����", wx.OK | wx.ICON_INFORMATION)
		fd.Destroy()
		event.Skip()
	
	def OnPreview( self, event ):
		# get current picture to display
		img = self.parent.getCurJpg()
		self.m_preview.Enable(False)
		wx.BeginBusyCursor()	
		self.DoEffects(img)
		wx.EndBusyCursor()
		self.m_preview.Enable(True)
		event.Skip()
		
	def OnReset( self, event ):
		self.parent.imgReload()
		self.parent.rubber.crop = None #������ѡ������
		event.Skip()
		
	def OnClose( self, event ):
		self.Show(False)
		event.Skip()
	
	def DoEffects( self, filename, outfilename = '', istemp = True ):
		# check if single mode
		if self.parent.singleMode and self.parent.im_preview:			
			self.cur_im = self.parent.im_preview			
		else: 	self.cur_im = Image.open(filename)
		
		# define current image
		self.cur_img = filename
		
		#
		btnPressedList = []
		[btnPressedList.append(btn) for btn in self.taskListBtns if btn.GetValue() ]		
			
		# ����û���������ˣ����û�����
		for btn in btnPressedList:
			if btn.Label not in self.m_checkList1.Items:
				self.m_checkList1.Append(btn.Label)
		#[self.m_checkList1.Append(btn.Label) for btn in btnPressedList if (btn.Label not in self.m_checkList1.Items)]

		# move image split to end if available
		index = self.m_checkList1.FindString(u"ͼƬ�ָ�")
		if index != wx.NOT_FOUND and index != self.m_checkList1.GetCount() - 1:
			self.m_checkList1.Delete(index)
			self.m_checkList1.Append(u"ͼƬ�ָ�")

		self.out_dir = outfilename if not istemp else None
			
		tasks = self.m_checkList1.Items
		try:
			if tasks:
				for task in tasks:
					self.taskMapper[task]()
				if istemp:	
					self.parent.imgPreview(self.cur_im, '.PNG') #ext)
				elif self.cur_im:
					ext = os.path.splitext(filename)[1]
					if ext.upper() in ['.JPG', '.JPEG']:
						self.cur_im.save(outfilename, 'JPEG', quality = 95, optimize = True, dpi = self.parent.current_dpi)
					else: self.cur_im.save(outfilename, dpi = self.parent.current_dpi)
		except: raise

	def MakeRound (self):
		r = self.m_spinCtrlNum.GetValue()
		if not self.m_checkBoxDefault.GetValue():
			self.cur_im = round_image(self.cur_im, radius = self.m_spinCtrlNum.GetValue())
		else:
			self.cur_im = round_image(self.cur_im) # use default 5
		#return img
	def MakeBackground (self):      
		method = self.m_comboMethod.GetValue() #u'ƫ��' #'Scale' #'Tile'
		if self.m_radioBtnCkColor.GetValue(): choice_index = 0
		else: choice_index = 1
		
		mark = self.pic
		bgColor = self.m_bgColourPicker.GetColour()
		h_offset, v_offset = int(self.m_textCtrlHbox.GetValue()), int(self.m_textCtrlVbox.GetValue())
		self.cur_im = background(self.cur_im, FILL_CHOICES[choice_index], mark, bgColor, horizontal_offset = h_offset, vertical_offset = v_offset, method = method,
				  horizontal_justification=self.m_comboTextHPos.GetValue(),
				  vertical_justification=self.m_comboTextVPos.GetValue())
	def MakeShadow (self):
		h_offset, v_offset = int(self.m_textCtrlShadowHBox.GetValue()), int(self.m_textCtrlShadowVBox.GetValue())
		self.cur_im = drop_shadow(self.cur_im, horizontal_offset = h_offset, vertical_offset = v_offset)
	
	def MakeRoll (self):
		h_offset, v_offset = int(self.m_textCtrlRollHBox.GetValue()), int(self.m_textCtrlRollVBox.GetValue())
		self.cur_im = self.cur_im.offset(h_offset, v_offset)
		
	def onMethodChanged(self, event):
		self.offsetGroupStatus()
		event.Skip()            
		
	def offsetGroupStatus(self):
		isOffsetMethod = (self.m_radioBtnCkPic.GetValue() and self.m_comboMethod.GetValue() == u'ƫ��')
		self.m_staticTextHtxt.Enable(isOffsetMethod)
		self.m_textCtrlHbox.Enable(isOffsetMethod)
		self.m_staticTextVtxt.Enable(isOffsetMethod)
		self.m_textCtrlVbox.Enable(isOffsetMethod)
		self.m_comboTextHPos.Enable(isOffsetMethod)
		self.m_comboTextVPos.Enable(isOffsetMethod)
		self.m_staticTextHPos.Enable(isOffsetMethod)
		self.m_staticTextVPos.Enable(isOffsetMethod)
	def MakeBorder(self):
		width = int(self.m_textCtrlBorderBox.GetValue())  # may throw error!
		opacity = self.m_sliderOpacity.GetValue()
		borderColor = self.m_borderColourPicker.GetColour()
		self.cur_im = border(self.cur_im, border_width = width, color= borderColor, opacity = opacity)
	def MakeReflection (self):
		depth = int(self.m_textCtrlDepthBox.GetValue())
		opacity = self.m_sliderReflectionOpacity.GetValue()
		reflectionBgColor = self.m_ReflectionBgColourPicker.GetColour()
		self.cur_im = reflect(self.cur_im, depth, 90, reflectionBgColor, opacity, 'BICUBIC')

	def MakeGrayscale(self):
		self.cur_im = self.cur_im.convert('L')
		
	def MakeEnhance(self):
		modeMapper = {u'��':ImageEnhance.Sharpness, u'ɫ��':ImageEnhance.Color, u'����':ImageEnhance.Brightness, u'�Աȶ�':ImageEnhance.Contrast}
		factor = 1.0 + self.m_spinCtrlFactor.GetValue() / 70.0  # allow range from 0 to 2.9
		mode = self.m_EnhanceBox.GetValue()
		
		enhancer = modeMapper[mode](self.cur_im)
		self.cur_im = enhancer.enhance(factor)
	
	def MakeFilter(self):
		filtermode = [ImageFilter.BLUR, ImageFilter.GaussianBlur, ImageFilter.EMBOSS, ImageFilter.CONTOUR,
				  ImageFilter.EDGE_ENHANCE, ImageFilter.SMOOTH, ImageFilter.MedianFilter, ImageFilter.SHARPEN,
				  ImageFilter.DETAIL]
		for checkbox in self.m_FilterBoxList:           
			if checkbox.GetValue():
				index = [i for i in range(len(self.m_FilterBoxList)) if self.m_FilterBoxList[i] == checkbox]
				if  self.cur_im.mode != '1':
					self.cur_im = self.cur_im.convert("RGBA")
					self.cur_im = self.cur_im.filter(filtermode[index.pop()])
				
	def MakeBlkWht(self):
		self.cur_im = self.cur_im.convert('1')
		
	def MakeRotate(self):
		# First check if user selected fixed angle?
		rotate_map = {0:Image.ROTATE_270, 1:Image.ROTATE_180, 2:Image.ROTATE_90}
		select = self.m_choiceFixed.Selection
		if self.m_radioBtnFixed.GetValue():
			self.cur_im = self.cur_im.transpose(rotate_map[select])
		else:
			# make it clockwise for positive angle
			angle = -int(self.m_spinCtrlAny.GetValue()) 
			self.cur_im = self.cur_im.rotate(angle, expand = 1)
		
	def MakeMirror(self):
		direction = self.m_MirrorBox.GetValue()
		self.cur_im = tile(self.cur_im, direction)
		
	def MakeInvert(self):
		amount = self.m_spinCtrlInvert.GetValue() + 50 # -50->0, 50 ->100
		self.cur_im = invert(self.cur_im, amount)
		
	def MakeMolten(self):
		self.cur_im = molten(self.cur_im)
	
	def MakeLighting(self):
		W, H = self.cur_im.size
		
		self.m_PosList = [u'����', u'����', u'����', u'����', u'����', u'����', u'����', u'����',u'����']
		self.m_PosList_Map = {u'����':(W/4,H/4), u'����':(W/4, H/2), u'����':(W/4, (H*3)/4),
				u'����':(W/2, H/4), u'����':(W/2,H/2), u'����': (W/2, (H*3)/4),
				u'����':((W*3)/4, H/4), u'����':((W*3)/4, H/2),u'����':((W*3)/4, (H*3)/4)}
		pos = self.m_comboTextLightingPos.GetValue()
		power = self.m_spinCtrlLighting.GetValue()
		self.cur_im = lighting(self.cur_im, power, self.m_PosList_Map[pos])
		
	def MakeSketch(self):
		threshold = self.m_spinCtrlSketch.GetValue()
		self.cur_im = sketch(self.cur_im, threshold)
		
	def MakeIce(self):
		self.cur_im = ice(self.cur_im)
		
	def ColorWithAlpha(self, color, a):
		r,g,b = color
		return (r,g,b,a)
		
	def MakePapercut(self):
		threshold = self.m_spinCtrlPapercut.GetValue()
		bg_color = self.m_PapercutBgColourPicker.GetColour() #(255, 255, 255, 0)
		fg_color = self.m_PapercutFgColourPicker.GetColour() #(255, 0, 0, 255)
		self.cur_im = paper_cut(self.cur_im, threshold, self.ColorWithAlpha(bg_color,0), self.ColorWithAlpha(fg_color, 255))

	def MakeDeYellow(self):
		if self.m_radioBoxDeYellow_1.GetValue(): method = 0
		else: method = 1
		self.cur_im = deYellow(self.cur_im, method)
		
	def MakeTan(self):
		self.cur_im = tan(self.cur_im)
		
	def MakeSwirl(self):
		degree = self.m_sliderSwirlOpacity.GetValue()
		self.cur_im = swirl(self.cur_im, degree)
		
	def MakeSpherize(self):
		self.cur_im = spherize(self.cur_im)
		
	def MakePuzzle(self):
		#print "start puzzle"
		if self.m_PuzzleSrcBox.GetValue() == self.m_PuzzleSrcModeList[0]:		
			im_list = [ self.cur_im.convert("RGBA") ]
		else:
			im_list = [ Image.open(i).convert("RGBA") for i in self.parent.picPaths ]

		x_qty = self.m_spinCtrlQtyX.GetValue()
		y_qty = self.m_spinCtrlQtyY.GetValue()
		is_column = self.m_PuzzleOrderBox.GetValue() == u'����'
		gap = self.m_spinCtrlGap.GetValue()
		allow_resize = self.m_checkBoxResize.GetValue()
		self.cur_im = ImgSalon(x_qty, y_qty, im_list, self.ColorWithAlpha(self.m_PuzzleBgColourPicker.GetColour(), 255), is_column, gap, allow_resize)
	
	def MakeFocus (self):
		edge_color = self.ColorWithAlpha (self.m_FocusEdgeColourPicker.GetColour(), 255)
		radius = min(self.cur_im.size[0], self.cur_im.size[1]) / 2
		radius_in = radius * self.m_spinCtrlRadiusIn.GetValue() /100 
		radius_out = radius * self.m_spinCtrlRadiusOut.GetValue() /100
		edge_transparent = self.m_checkBoxFocusEdge.IsChecked()
		self.cur_im = edge_blur(edge_color, self.cur_im, radius_in, radius_out, edge_transparent)
		
	def MakeCornerFold (self):
		edge = self.m_spinCtrlEdge.GetValue()
		white_percent = self.m_spinCtrlWhitePercent.GetValue()
		pos = self.m_FoldPosBox.GetValue()
		self.cur_im = EdgeFold(self.cur_im, edge, white_percent, pos)
		
	def MakeTear (self):
		direction = self.m_TearPosBox.GetValue()
		pos = self.m_spinCtrlTear.GetValue()
		isCut = self.m_checkBoxIsCut.GetValue()
		self.cur_im = Tear(self.cur_im, pos, direction, isCut)
		
	def MakeFade (self):
		pos = self.m_spinCtrlFadePos.GetValue()
		mode = self.m_FadeModeBox.GetValue()
		color = self.ColorWithAlpha (self.m_FadeColourPicker.GetColour(), 255)
		self.cur_im = Fade(self.cur_im, pos, mode, color)
	def MakeMix (self):
		factor = self.m_spinCtrlMixFactor.GetValue()
		color = self.getColor(self.ctrlColor)
		self.cur_im = MixColor(self.cur_im, factor, color)
		
	def AddMask (self):
		if self.mask:
			bg_color = self.getColor(self.ctrlMaskBgColor)
			self.cur_im = AddMask(self.cur_im, self.mask, bg_color)
			
	def MakeBlend (self):
		if self.img2:
			alpha = '%.2f' %  (int(self.m_spinCtrlAlpha.GetValue()) /100.0)
			self.cur_im = blend (self.cur_im, self.img2, float(alpha))
			
	def MakeSplit (self):
		rows, cols = self.m_spinCtrlRow.GetValue(), self.m_spinCtrlCol.GetValue()
		isPx, isPy = self.m_choicePixelX.GetSelection() == 1, self.m_choicePixelY.GetSelection() == 1
		w, h = self.cur_im.size
		
		# check if valid values
		if rows > w: rows = w
		if cols > h: cols = h
		
		img = self.parent.getCurJpg()
		line_color = self.m_LineColourPicker.GetColour().Get(includeAlpha=True)

		out_dir = os.path.splitext(self.out_dir)[0] if self.out_dir else None
		
		self.cur_im = split (self.cur_im, rows, cols, line_color, isPx, isPy, out_dir )
		
	def SliceGone (self):
		horizontal = self.m_checkBoxSliceStyle.GetValue() == u"����"
		if self.parent.rubber.crop:
			x1,y1,x2,y2 = self.parent.rubber.crop
			if not horizontal:
				self.cur_im = slice_gone(self.cur_im, y1, y2, False)
			else: 
				self.cur_im = slice_gone(self.cur_im, x1, x2, True)
			
	def MakeWaterMark (self):
		#img = self.parent.getCurJpg()
		img = self.cur_img
		self.cur_im = self.parent.txtDlg.DoWaterMark(img, self.cur_im)
		
	def OnMixColorChanged (self, event):
		self.colorPicker2Ctrl(self.m_MixColourPicker, self.ctrlColor)
		event.Skip()
		
	def OnMaskBgColorChanged (self, event):
		self.colorPicker2Ctrl(self.m_MaskBgColorPicker, self.ctrlMaskBgColor)
		event.Skip()
		
	def OnMaskBgTxtChanged (self, event):
		self.ctrl2ColorPicker(self.ctrlMaskBgColor, self.m_MaskBgColorPicker)
		event.Skip()
		
	def OnMixTxtChanged (self, event):
		self.ctrl2ColorPicker(self.ctrlColor, self.m_MixColourPicker)
		event.Skip()
	
	def OnOldTxtChanged (self, event):
		self.ctrl2ColorPicker(self.ctrlOldColor, self.m_OldColourPicker)
		event.Skip()

	def OnOldTxtMaxChanged (self, event):
		self.ctrl2ColorPicker(self.ctrlOldColorMax, self.m_OldColourMaxPicker)
		event.Skip()
		
	def OnNewTxtChanged (self, event):
		self.ctrl2ColorPicker(self.ctrlNewColor, self.m_NewColourPicker)
		event.Skip()
		
	def OnOldColorChanged (self, event):
		self.colorPicker2Ctrl(self.m_OldColourPicker, self.ctrlOldColor)
		event.Skip()
		
	def OnOldColorMaxChanged (self, event):
		self.colorPicker2Ctrl(self.m_OldColourMaxPicker, self.ctrlOldColorMax)
		event.Skip()
		
	def OnNewColorChanged (self, event):
		self.colorPicker2Ctrl(self.m_NewColourPicker, self.ctrlNewColor)
		event.Skip()
	def OnShearTxtChanged (self, event):
		self.ctrl2ColorPicker(self.ctrlShearBkColor, self.m_ShearColorPicker)
		event.Skip()
		
	def OnShearColorChanged (self, event):
		self.colorPicker2Ctrl(self.m_ShearColorPicker, self.ctrlShearBkColor)
		event.Skip()
		
	def colorPicker2Ctrl (self, colorPicker, txtCtrl):
		r, g, b= colorPicker.GetColour()
		color = r, g, b, 255
		txtCtrl.SetValue(str(color))
	
	def ctrl2ColorPicker (self, txtCtrl, colorPicker):
		color = txtCtrl.GetValue()
		try:
			if '#' not in color: color = eval(color)
			colorPicker.SetColour(color)
			colorPicker.Refresh()
		except: pass
	def getColor(self, txtCtrl):
		color_str = txtCtrl.GetValue()
		if '#' not in color_str: color = eval(color_str)
		else:
			color = wx.Colour()
			color.SetFromString(color_str)
			color = color.Get(True)
		return color

	def MakeReplace (self):		
		color_old = self.getColor(self.ctrlOldColor)
		color_old_max = self.getColor (self.ctrlOldColorMax)
		color_new = self.getColor(self.ctrlNewColor)
		flag = self.m_ColorCompareBox.GetValue()
		w, h = self.cur_im.size
		im_array = np.array(self.cur_im.convert("RGBA"))
		if self.parent.rubber.crop:
			x1,y1,x2,y2 = self.parent.rubber.crop
			region_array = im_array[y1:y2, x1:x2]
			if self.m_checkBoxReplaceStyle.Selection == 0:	
				region_array_new = Replace(region_array, color_old, color_new, flag, color_old_max)								
			else:				
				region = newPil(x2-x1, y2-y1, fill = color_new)				
				region_array_new = np.array(region).astype(np.uint8)			
	
			im_array[ y1:y2, x1:x2] = region_array_new			
			self.cur_im = Image.fromarray(im_array)			
		else:
			if self.m_checkBoxReplaceStyle.Selection == 0:				
				self.cur_im = Image.fromarray(Replace(im_array, color_old, color_new, flag, color_old_max))
			else: # create a numpy image with color_new
				self.cur_im = newPil(w,h,color_new)
		
	def MakeAutoBrightness (self):
		new_brightness = self.m_spinCtrlNewBrightness.GetValue()
		self.cur_im = equalize(self.cur_im, new_brightness )
		
	def MakeShear (self):
		x_degree = self.m_spinCtrlXdegree.GetValue()
		y_degree = self.m_spinCtrlYdegree.GetValue()
		bk_color = self.getColor(self.ctrlShearBkColor)
		self.cur_im = Shear(self.cur_im, x_degree, y_degree, bk_color)
	
	def MakeMosaic (self):
		mosaic_size = self.m_spinCtrlMosaicSize.GetValue()
		if self.parent.rubber.crop:
			x0, y0, x1, y1 = self.parent.rubber.crop
		else:
			x0, y0, x1, y1 = 0, 0, self.cur_im.size[0], self.cur_im.size[1]
		self.cur_im = mosaic(self.cur_im, x0, y0, x1, y1, mosaic_size)		
		
	#---------------------------------------------------------------------                
	def doTxtDlg (self):
		"""
		Start Water Mark GUI
		"""
		pub.sendMessage("set status", msg = u"С��ʾ��ʹ��ͼƬˮӡ���ܣ���������������Լ���ӡ�ºǡ�")
		# only popup font dlg after jpg is loaded
		if not self.parent.txtDlg:
			self.parent.txtDlg = MyTextDialog(self.parent)
			self.parent.setFontDlgPos(self.parent.txtDlg)                                
		else:
			self.parent.setFontDlgPos(self.parent.txtDlg)
			# get saved pin state
			self.parent.enablePin = (self.parent.txtDlg.textRepeatCtrl.GetValue() == u'һ�� -> �����ָ��λ��')                
##################################################################
class MyTextDialog(wx.Dialog):
        """"""
        #-------------------------------------------------------------
        def __init__(self, parent):                
                """�����ʼ������"""

                self.parent = parent
                
                #��������
                self.Base = 40 # old = 20
                #ˮӡ��
                self.text = self.parent.water_text
                self.txtEnabled = True                
               
                #ϵͳ���弯�ֵ�
                self.systemFontDict = self._GetSystemFontDict()
                
                #������ʽ
                self.fontStyle = None
                
                #�����ļ�
                self.fontFile = "SIMSUN.TTC"
                
                #����·��
                self.fontSrc = os.path.join(os.environ['windir'], "fonts") ##"C:\\WINDOWS\\Fonts\\"
                #�����С
                self.fontSize = 20
                #������ɫ,Ĭ�Ϻ�ɫ
                self.fontColour = (0,0,0) # 
                #͸����
                self.trans = self.parent.text_trans
                #ȡ������ɫ = ��ɫ + ͸����
                #self._GetFinalColourWithAlpha()
                #ˮӡλ��,Ĭ��Ϊ����
                self.fontPosition = 9 #the default
                
                # Do not initialize your font here; Get the font chosen by the user
                self.curFont =  wx.Font(self.fontSize, 70, 90, 90, False, wx.EmptyString )
                self.fName = unicode("����", "cp936")
                self.LABEL_FONT = wx.Font(9, wx.SWISS, wx.NORMAL, wx.NORMAL, False,'Tahoma')

                wx.Dialog.__init__(self, parent, title='������ֻ�ͼƬˮӡ', size=(600, 820), style = wx.DEFAULT_DIALOG_STYLE |wx.MINIMIZE_BOX )
                self.SetIcon(self.parent.midi)                
                
                self.sizer = wx.BoxSizer(wx.VERTICAL)
                self.sizer.Add(wx.StaticLine(self, wx.ID_ANY), 0, wx.ALL|wx.EXPAND, 0)
                #self.watermarkSizer = wx.BoxSizer(wx.HORIZONTAL)
                
                #ˮӡ����
                staticChoice = wx.StaticBox (self, wx.ID_ANY, u"ˮӡ��Դ")
                sizerChoice = wx.StaticBoxSizer(staticChoice, wx.HORIZONTAL)
                
                fgSizerSrc = wx.FlexGridSizer( 2, 2, 0, 50 )
                fgSizerSrc.SetFlexibleDirection(wx.BOTH)
                fgSizerSrc.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
                
                bSizerPicName = wx.BoxSizer(wx.VERTICAL)
                self.m_radioPicName = wx.RadioButton( self, wx.ID_ANY, u"ͼƬ����", wx.DefaultPosition, wx.DefaultSize, wx.RB_GROUP)
                bSizerPicName.Add( self.m_radioPicName, 0, wx.ALL, 5 )
                fgSizerSrc.Add( bSizerPicName, 1, wx.EXPAND, 5 )
                self.m_radioPicName.Bind(wx.EVT_RADIOBUTTON, self.OnRadioBtns)
                
                # ͼƬˮӡѡ��  
                bSizerPic = wx.BoxSizer( wx.HORIZONTAL )
                self.m_radioPic = wx.RadioButton( self, wx.ID_ANY, u"ͼƬˮӡ", wx.DefaultPosition, wx.DefaultSize, 0 )
                bSizerPic.Add( self.m_radioPic, 0, wx.ALL, 5 )
                self.m_radioPic.Bind(wx.EVT_RADIOBUTTON, self.OnRadioBtns)
                          
                self.m_textCtrlPath = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, (300, -1), wx.TE_READONLY )
                bSizerPic.Add( self.m_textCtrlPath, 0, wx.ALL, 5 )
                self.pic = None
                self.m_btnBrowse = wx.Button( self, wx.ID_ANY, u"...", wx.DefaultPosition, (-1, 21), wx.BU_EXACTFIT )
                self.m_btnBrowse.Bind(wx.EVT_BUTTON, self.onPicBrowse)
                bSizerPic.Add( self.m_btnBrowse, 0, wx.ALL, 5 )
                fgSizerSrc.Add( bSizerPic, 1, wx.ALL, 0 )
                
                bSizerUniText = wx.BoxSizer(wx.VERTICAL)
                self.m_radioTxt = wx.RadioButton( self, wx.ID_ANY, u"����ˮӡ", wx.DefaultPosition, wx.DefaultSize, 0 )                
                self.m_radioTxt.Bind(wx.EVT_RADIOBUTTON, self.OnRadioBtns)
                bSizerUniText.Add( self.m_radioTxt, 0, wx.ALL, 5 )
                fgSizerSrc.Add( bSizerUniText, 1, wx.EXPAND, 5 )
                
                #����ˮӡ�б�
                bSizerFileList = wx.BoxSizer( wx.HORIZONTAL )
                self.m_radioFileList = wx.RadioButton( self, wx.ID_ANY, u"�����ļ�", wx.DefaultPosition, wx.DefaultSize, 0 )
                bSizerFileList.Add( self.m_radioFileList, 0, wx.ALL, 5 )
                self.m_radioFileList.Bind(wx.EVT_RADIOBUTTON, self.OnRadioBtns)
                
                self.m_fileListCtrlPath = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, (300, -1), wx.TE_READONLY )
                bSizerFileList.Add( self.m_fileListCtrlPath, 0, wx.ALL, 5 )
                
                self.listFile = None
                self.excel_pic_names = None
                
                self.m_btnBrowseFileList = wx.Button( self, wx.ID_ANY, u"...", wx.DefaultPosition, (-1, 21), wx.BU_EXACTFIT )
                self.m_btnBrowseFileList.Bind(wx.EVT_BUTTON, self.onFileListBrowse)
                bSizerFileList.Add( self.m_btnBrowseFileList, 0, wx.ALL, 5 )
                fgSizerSrc.Add( bSizerFileList, 1, wx.EXPAND, 0 )
                
                sizerChoice.Add(fgSizerSrc, 0, wx.EXPAND, 5 )
                self.sizer.Add(sizerChoice, 0, wx.LEFT | wx.TOP | wx.RIGHT | wx.EXPAND, 13)
                self.m_radioTxt.SetValue(True)
                
                ##
                self.sizer.AddSpacer( ( -1, 5))
                ###
                self.staticTxtMark = wx.StaticBox( self, wx.ID_ANY, u"����ˮӡ" )
                self.bSizerSrcWaterMark = wx.StaticBoxSizer( self.staticTxtMark, wx.VERTICAL )
                
                self.waterText = wx.TextCtrl(id=wx.ID_ANY, name='waterText',
                      parent=self, pos=wx.Point(88, 64), size=wx.Size(556, 300),
                      style=wx.TE_MULTILINE,
                      value=self.text)
                self.waterText.SetMaxLength(400)
                self.waterText.SetToolTipString("���200������")
                self.waterText.Bind(wx.EVT_TEXT, self.OnWaterText)
                self.bSizerSrcWaterMark.Add(self.waterText, 0, wx.ALL | wx.CENTER, 2)
                
                # ��ǰ����״̬
                self.statusSizer = wx.BoxSizer(wx.HORIZONTAL)
                font_status = self.GetFontStatus()                
                self.fontStatusLabel = wx.StaticText(self, label = font_status, size = (332, -1))
                self.fontStatusLabel.SetFont(self.LABEL_FONT)
                self.fontStatusLabel.SetForegroundColour(self.fontColour)
                self.statusSizer.Add(self.fontStatusLabel, 0, wx.TOP | wx.LEFT, 5)
                
                #ѡ������                
                self.chooseFont = wx.Button(self, label = "���ĵ�ǰ����")
                self.chooseFont.SetFont(self.LABEL_FONT)
                self.Bind(wx.EVT_BUTTON, self.OnChooseFont, self.chooseFont)
                self.statusSizer.Add(self.chooseFont, 0, wx.TOP |wx.LEFT | wx.BOTTOM, 5)
                               
                # ˮӡλ��SIZER
                self.waterPanelSizer = wx.BoxSizer(wx.HORIZONTAL)
                
                #ˮӡλ�����				
                self.waterPosition = wx.StaticBox(self, label='ˮӡλ��', size=wx.Size(220, 180))
                self.waterPosition.SetFont(self.LABEL_FONT)
                self.btnGroupBoxSizer = wx.StaticBoxSizer(self.waterPosition, wx.VERTICAL)
                #self.waterPanelSizer.Add(self.waterPosition, 0, wx.LEFT, 8)
                
                gSizerBtnGroup = wx.GridSizer( 3, 3, 0, 0 )
                
                #ˮӡλ��9��
                self.LeftTop = wx.RadioButton(self, label='����', pos = wx.DefaultPosition, style = wx.RB_GROUP)
                self.LeftTop.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.CenterTop = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.CenterTop.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.RightTop = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.RightTop.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.LeftCenter = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.LeftCenter.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.Center = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.Center.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.RightCenter = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.RightCenter.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.LeftBotm = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.LeftBotm.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.CenterBotm = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)
                self.CenterBotm.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                self.RightBotm = wx.RadioButton(self, label='����', pos = wx.DefaultPosition)                
                self.RightBotm.Bind(wx.EVT_RADIOBUTTON, self.OnRadioButton)
                
                self.NineBtns = [self.LeftTop, self.CenterTop, self.RightTop, self.LeftCenter, self.Center, self.RightCenter,
                            self.LeftBotm, self.CenterBotm, self.RightBotm]
                for btn in self.NineBtns:
                        btn.SetFont(self.LABEL_FONT)
                        gSizerBtnGroup.Add( btn, 0, wx.ALL, 15 )
                self.btnGroupBoxSizer.Add( gSizerBtnGroup, 1, wx.EXPAND, 0 )
                self.waterPanelSizer.Add(self.btnGroupBoxSizer, 1, wx.ALL |wx.EXPAND, 15)
           
                self.fontEffectSizer = wx.BoxSizer(wx.VERTICAL)
                
                #����Ч������
                self.textRepeat = wx.StaticText(self, label='ˮӡ�ظ�����:')
                self.textRepeat.SetFont(self.LABEL_FONT)
                self.repeatList = [u'һ�� -> �����ˮӡλ��', u'һ�� -> �����ָ��λ��', u'���']#, u'���']                
                # �����ظ������ַ���
                if self.parent.water_text_repeat in [self.repeatList[0], self.repeatList[2]]: value = self.parent.water_text_repeat
                else: value = self.repeatList[0]
                
                self.textRepeatCtrl = wx.ComboBox(self, value = value, choices = self.repeatList, size = (160, -1), style = wx.CB_READONLY)
                self.Bind(wx.EVT_COMBOBOX, self.OnSelect, self.textRepeatCtrl)
                
                self.fontEffectSizer.Add((10, -1))
                
                #��͸���� 
                self.pickTrans = wx.StaticText(self, label='��͸����(0~100):')
                self.pickTrans.SetFont(self.LABEL_FONT)
                self.transCtrl = wx.Slider(self, maxValue=100, minValue=0, size = (160, -1),
                        style=wx.SL_HORIZONTAL|wx.SL_LABELS, value = self.parent.text_trans)
                
                # ��ת
                self.textRotate = wx.StaticText(self, label = '��ת�Ƕ�(0~360):')
                self.textRotate.SetFont(self.LABEL_FONT)
                value = self.parent.water_text_angle
                self.textRotateCtrl = wx.Slider(self, maxValue=360, minValue=0, size = (160, -1),
                        style=wx.SL_HORIZONTAL|wx.SL_LABELS, value = value)
                
                self.m_btnBrowse.Enable(False)
                self.m_btnBrowseFileList.Enable(False)
                
                self.fontEffectSizer.Add(self.textRepeat, 0, wx.ALL | wx.TOP, 2)
                self.fontEffectSizer.Add(self.textRepeatCtrl, 0, wx.ALL | wx.TOP, 2)
                self.fontEffectSizer.Add((-1, 10))
                self.fontEffectSizer.Add(self.pickTrans, 0, wx.ALL | wx.TOP, 2)
                self.fontEffectSizer.Add(self.transCtrl , 0, wx.ALL | wx.TOP , 2)
                self.fontEffectSizer.Add(self.textRotate , 0, wx.ALL | wx.TOP , 2)
                self.fontEffectSizer.Add(self.textRotateCtrl , 0, wx.ALL | wx.TOP , 2)
                
                self.waterPanelSizer.Add(self.fontEffectSizer, 1, wx.LEFT | wx.TOP, 10)
                
                # add a slide factor
                self.bSizerFaq = wx.BoxSizer( wx.VERTICAL )
        
                self.m_staticTextFaq = wx.StaticText( self, wx.ID_ANY, u"Ƶ�ʵ���", wx.DefaultPosition, wx.DefaultSize, 0 )
                self.m_staticTextFaq.Wrap( -1 )
                self.bSizerFaq.Add( self.m_staticTextFaq, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
                
                self.m_sliderFaq = wx.Slider( self, wx.ID_ANY, 5, 1, 10, wx.DefaultPosition, wx.DefaultSize, wx.SL_LABELS|wx.SL_VERTICAL )
                self.bSizerFaq.Add( self.m_sliderFaq, 0, wx.ALIGN_CENTER|wx.ALL, 5 )
                
                self.waterPanelSizer.Add(self.bSizerFaq, 1, wx.TOP, 15)
                # finish the slide faq

                self.bSizerSrcWaterMark.Add(self.statusSizer, 0, wx.LEFT, 3)       
                self.sizer.Add(self.bSizerSrcWaterMark, 0, wx.ALL | wx.CENTER, 1)                                   
                self.sizer.Add((-1, 10))
                self.sizer.Add(self.waterPanelSizer, 0, wx.ALL | wx.EXPAND, 0)
                
                # add a separator
                self.sizer.Add(wx.StaticLine(self, wx.ID_ANY),
                                   0, wx.ALL|wx.EXPAND, 5)
                self.btnSizer = wx.BoxSizer(wx.HORIZONTAL)
                
                # toggle between
                self.ToggleRepeatFeq() 
                
                okBtn = wx.BitmapButton( self, wx.ID_ANY, self.parent.preview_ico, wx.DefaultPosition, wx.Size(48, 48), wx.NO_BORDER |wx.BU_EXACTFIT  )
                okBtn.Bind(wx.EVT_BUTTON, self.OnOK)
                self.btnSizer.Add(okBtn, 0, wx.ALL | wx.Center, 6)
                okBtn.SetToolTipString(u'Ԥ��Ч��')
                
                resetBtn = wx.BitmapButton( self, wx.ID_ANY, self.parent.refresh_ico, wx.DefaultPosition, wx.Size(48, 48), wx.NO_BORDER|wx.BU_EXACTFIT  )
                resetBtn.Bind(wx.EVT_BUTTON, self.OnReset)
                self.btnSizer.Add(resetBtn, 0, wx.ALL | wx.Center, 6)
                resetBtn.SetToolTipString(u'���¼���ͼƬ')
                
                cancelBtn = wx.BitmapButton( self, wx.ID_ANY, self.parent.close32_ico, wx.DefaultPosition, wx.Size(48, 48), wx.NO_BORDER|wx.BU_EXACTFIT  )
                cancelBtn.Bind(wx.EVT_BUTTON, self.OnCancel)    
                self.btnSizer.Add(cancelBtn, 0, wx.TOP | wx.RIGHT, 6)
                cancelBtn.SetToolTipString(u'�ر�')

                self.sizer.Add(self.btnSizer, 0, wx.TOP | wx.ALIGN_RIGHT, 6)                
                self.SetSizer(self.sizer)
                
                self.UpdateText()                
                self.RightBotm.SetValue(True) # Ĭ��ѡ��
                self.m_radioTxt.SetValue(True)
                
        def OnTxtSrc (self, event):
                event.Skip()
                
        def OnTxtSrcChoice (self, event):
                event.Skip()
                
        def OnRadioBtns (self, event):
                self.txtEnabled = self.m_radioTxt.GetValue()
                # process pic buttons
                self.m_btnBrowse.Enable( self.m_radioPic.GetValue())
                if self.m_btnBrowse.IsEnabled(): self.m_btnBrowse.SetFocus()
                # process text buttons
                self.m_btnBrowseFileList.Enable( self.m_radioFileList.GetValue())
                if self.m_btnBrowseFileList.IsEnabled(): self.m_btnBrowseFileList.SetFocus()
                
                self.ToggleWatermarkSrc(self.txtEnabled)
                self.setRotateCtrl()
                event.Skip()
        #---------------------------------------------------------
               
        def onPicBrowse(self, event):
                wildcard = "�����ļ� (*.*)|*.*|JPEG ͼƬ (*.jpg)|*.jpg|BMP ͼƬ (*.bmp)|*.bmp|PNG ͼƬ (*.png)|*.png|" +\
                        "GIF ͼƬ (*.gif)|*.gif|TIF ͼƬ (*.tif; *.tiff)|*.tif; *.tiff|ICO ͼ�� (*.ico)|*.ico" 
                        
                fd = wx.FileDialog(self, u'��ѡ��һ��ˮӡͼƬ', os.getcwd(), '', wildcard, wx.FD_CHANGE_DIR | wx.OPEN)
                if fd.ShowModal() == wx.ID_OK:
                        self.pic = fd.GetPath()
                        self.m_textCtrlPath.SetValue(self.pic)
                        self.m_textCtrlPath.SetToolTip(wx.ToolTip(self.pic))
                fd.Destroy()
                event.Skip()
                
        def onFileListBrowse (self, event):
                wildcard = "�ı��ļ� (*.txt)|*.txt|Excel�ļ� (*.xlsx)|*.xlsx|�����ļ� (*.*)|*.*"                 
                if self.m_fileListCtrlPath.GetValue() != "": initDir = os.path.dirname(self.m_fileListCtrlPath.GetValue())
                else: initDir = os.getcwd()
                
                fd = wx.FileDialog(self, u'��ѡ���Ӧ���б��ļ�', initDir , '', wildcard, wx.FD_CHANGE_DIR | wx.OPEN)
                
                if fd.ShowModal() == wx.ID_OK:
                        filein = fd.GetPath()
                        if os.path.isfile(filein): self.ReadExternalFile(filein)   # check if file exitsf                        
                fd.Destroy()
                event.Skip()
                
        def ReadExternalFile(self, filein):
                
                # reset lists
                self.excel_pic_names = []
                self.listFile = []
                try:
                        wb = load_workbook(filein)
                        sheet = wb.get_active_sheet()
                        rows =  sheet.get_highest_row()
                        cols = sheet.get_highest_column()

                        for row in xrange(rows):
                                v = sheet.cell(row=row + 1, column=1).value
                                line = sheet.cell(row=row + 1, column=2).value
                                self.excel_pic_names.append(v.lower())
                                self.listFile.append(line)
                        self.CheckListFiles(rows, filein)
                        
                        
                except InvalidFileException:
                        with open(filein, 'rb') as fileIn:
                                lines = unicode(fileIn.read(),'cp936')
                                rows = len(lines.split("\r\n"))
                                self.listFile = lines.split("\r\n")
                                self.CheckListFiles(rows, filein)
                                
                                
        def CheckListFiles(self, rows, filein):
                if  rows == len(self.parent.picPaths):                        
                        if True in [ (len(j) > 200 ) for j in self.listFile]:
                                wx.MessageBox("�б����е�ˮӡ�����������ƣ������ļ���", "С����", wx.OK | wx.ICON_INFORMATION)
                                self.listFile = None
                                self.m_fileListCtrlPath.Clear()
                        else:
                                self.m_fileListCtrlPath.SetValue(filein)
                                self.m_fileListCtrlPath.SetToolTip(wx.ToolTip(filein))                                               
                else:
                        error = 'ͼƬ����( %s )�������б�( %s )��ƥ�䣡�����ԡ�' % (len(self.parent.picPaths), rows)
                        self.listFile = None
                        self.m_fileListCtrlPath.Clear()
                        wx.MessageBox(error, "С����", wx.OK | wx.ICON_INFORMATION)
                        
                        
                
        #---------------------------------------------------------
        def ToggleWatermarkSrc(self, enableTxt = True):
                #self.m_textCtrlPath.Enable(not enableTxt)
                #self.m_btnBrowse.Enable(not enableTxt)                
                self.staticTxtMark.Enable(not self.m_radioPic.GetValue())
                self.waterText.Enable(enableTxt)
                self.fontStatusLabel.Enable(not self.m_radioPic.GetValue())
                self.chooseFont.Enable(not self.m_radioPic.GetValue()) 
        #----------------------------------------------------------
        def GetFontStatus(self):
                return u"��ǰ����: " + self.curFont.GetNativeFontInfoUserDesc(). \
                replace("windows-936", u"����").replace("windows-1252", u"��������")
        #----------------------------------------------------------
        def OnSelect(self, event):
                # reload image if not single mode
                if not self.parent.singleMode: self.parent.imgReload()
                self.parent.water_text_repeat = self.textRepeatCtrl.GetValue()
                
                if self.textRepeatCtrl.GetValue() == u'һ�� -> �����ָ��λ��':
                        if self.parent.show_custom_warning == 'YES':
                                wx.MessageBox(u"��������������ˮӡ��λ�á�", u"�Զ���ˮӡλ��", wx.OK | wx.ICON_INFORMATION)
                                self.parent.show_custom_warning = 'NO'
                        else: pub.sendMessage("set status", msg = u"��������������ˮӡ��λ�á�")
                        # enable crop when the control is enbaled
                        if not self.parent.rubber: self.parent.rubber = wxPyRubberBander(self.parent.imageCtrl)
                        
                # 0 = no repeat; 1 = repeat
                self.ToggleRepeatFeq()
                self.parent.startPoint = None                
                event.Skip()
                
        ##----------------------------------------------------------
        def ToggleRepeatFeq(self):
                # define enable or not
                repeat_once = (u'һ��' in self.textRepeatCtrl.GetValue())
                self.parent.enablePin = (self.textRepeatCtrl.GetValue() == u'һ�� -> �����ָ��λ��')
                self.m_staticTextFaq.Enable(not repeat_once)
                self.m_sliderFaq.Enable(not repeat_once)
                posFlag = (self.textRepeatCtrl.GetValue() == u'һ�� -> �����ˮӡλ��')
                self.waterPosition.Enable(posFlag)
                for radioBtn in self.NineBtns: radioBtn.Enable(posFlag) 
                self.setRotateCtrl()
                # ��ʼ�� google drop
                if self.parent.enablePin:
                        if not self.parent.rubber:
                                self.parent.rubber = wxPyRubberBander(self.parent.imageCtrl)
        ##----------------------------------------------------------
        def setRotateCtrl(self):
                flag = u'һ��' in self.textRepeatCtrl.GetValue() #and (self.txtEnabled or self.m_radioFileList.GetValue()))
                self.textRotateCtrl.Enable(not flag )
                self.textRotate.Enable(not flag)
        #----------------------------------------------------------
        def OnChooseFont(self, event):
                """
                ѡ������ʱ,�ı�Ԥ�����������ʽ,������ֵ��ȫ�ֱ���fontFile
                """
                fontData = wx.FontData()
                dlg = wx.FontDialog(self, fontData)
                fontData = dlg.GetFontData()
                fontData.SetInitialFont(self.curFont)
                fontData.SetColour(self.fontColour)
                #fontData.SetAllowSymbols(False)                
                #fontData.EnableEffects(False)                 
                
                if dlg.ShowModal() == wx.ID_OK:
                        fontData = dlg.GetFontData()
                        #ȡ����ѡ����
                        self.curFont = fontData.GetChosenFont()
                        #ȡ��������ɫ                        
                        self.fontColour = fontData.GetColour()
                        
                        #���������С
                        self.fontSize = self.curFont.GetPointSize()
                        #ȡ��ʽ��ϵͳ����TTF��                    
                        self.fName = self.curFont.GetFaceName()
                        #ȡ������ʽ
                        self.fontStyle  = self.curFont.GetStyle()
                        
                        # set font status label                        
                        self.fontStatusLabel.SetLabel(self.GetFontStatus())
                        self.fontStatusLabel.SetForegroundColour(self.fontColour)
                        #self.fontStatusLabel.SetFont(self.curFont)

                        if self.fName in [unicode(font,"cp936") for font in ["����", "����-PUA","������"]]:
                                self.fName = unicode("SimSun & NSimSun", "cp936")
                        elif self.fName == unicode("����","cp936"):
                                self.fName = unicode("FangSong", "cp936")
                        elif self.fName == unicode("����","cp936"):
                                self.fName = unicode("SimHei", "cp936")
                        elif self.fName == unicode("����","cp936"):
                                self.fName = unicode("KaiTi", "cp936")
                        elif self.fName == unicode("΢���ź�","cp936"):
                                self.fName = unicode("Microsoft YaHei", "cp936")
                                
                        # process font name starts with "@"
                        elif self.fName in [unicode(font, "cp936") for font in ["@����", "@������", "@����"]]:
                                self.fName = unicode("SimSun-ExtB", "cp936")
                        elif self.fName == unicode("@����", "cp936"):
                                self.fName = unicode("DFKai-SB", "cp936")
                        elif self.fName == unicode("@����", "cp936"):
                                self.fName = unicode("Microsoft JhengHei & Microsoft JhengHei UI", "cp936")
                        elif self.fName == unicode("@΢���ź�", "cp936"):
                                self.fName = unicode("Microsoft YaHei & Microsoft YaHei UI", "cp936")
                        elif self.fName.startswith(unicode("@", "cp936")):
                                self.fName = unicode("MingLiU & PMingLiU & MingLiU_HKSCS", "cp936")
                        #if self.fName==unicode("����","cp936") or self.fName==unicode("����-PUA","cp936") or self.fName==unicode("������","cp936"):
                        #        self.fName = unicode("���� & ������","cp936")
                        try:
                                self.fontFile = self.systemFontDict[self.fName] 
                        except:
                                # upon exception, use default font                          
                                self.fName = unicode("SimSun & NSimSun", "cp936")
                                self.fontFile = self.systemFontDict[self.fName]
                        finally:
                                self.UpdateText()
                                self.waterText.Refresh()
                dlg.Destroy()
                event.Skip()
        #----------------------------------------------------------
        def UpdateText(self):                
                #self.curFont = wx.Font(self.fontSize, self.family, self.fontFile, self.weight, self.isUnderlined, self.fName)
                self.waterText.SetFont(self.curFont)
                self.waterText.SetOwnForegroundColour(self.fontColour) 
                #self.waterText.SetStrikethrough(self.strikethrough)

        #-----------------------------------------------------------
        def OnWaterText(self, event):
                """
                ���޸�ˮӡ���ֿ�ʱ,ͬ������text��ֵ
                """
                self.text = self.waterText.GetValue()
                self.parent.water_text = self.text
                event.Skip()
         #------------------------------------------------------------
        def DoWaterMark(self, filename, im):
                """"""
                return self._im_mark(filename, False, im)                
        
        #-----------------------------------------------------------
        def _GetSystemFontDict(self):
                """
                ȡϵͳ�����弯,����һ���ֵ����͵Ķ���,�ڳ�ʼ��ʱ������.
                """
                
                hkey = _winreg.OpenKey(_winreg.HKEY_LOCAL_MACHINE,
                                       "Software\\Microsoft\\Windows NT\\CurrentVersion\\Fonts") 
                nSubKey, nSubValue, lastModify = _winreg.QueryInfoKey(hkey)
                d = {}
                for i in range(nSubValue): 
                        valueName, valueObj,valueType = _winreg.EnumValue(hkey, i)
                        key = valueName.split(" (")[0]
                        d[unicode(key,"cp936")] = valueObj

                _winreg.CloseKey(hkey)
                return d
        
        def pic_watermark(self, isRepeat, ANGLE, percentage, filename, faq, im = None):
                if im:
                        im = im.convert("RGBA")
                        img = self.parent.wand2pil(im, 'PNG')
                # check if single mode
                elif self.parent.singleMode and self.parent.im_preview:
                        img = self.parent.wand2pil(self.parent.im_preview, 'PNG')
                else:   img = self.parent.readWandImage(filename)
                
                try:
                        imgTop = self.parent.readWandImage(self.pic)
                        imgTop.rotate(ANGLE) 
                        textwidth, textheight = imgTop.size
                        if (textwidth > self.parent.W or textheight > self.parent.H):
                                wx.MessageBox("��ѡ���ͼƬ̫���ˣ�����С�ߴ�����ԡ�", "С����", wx.OK | wx.ICON_INFORMATION)
                                return
                except:  return
                
                isPosSet, xy = self.isMousePos(textwidth, textheight,img.size[0], img.size[1])
                if not isPosSet: return
                
                #draw.setfont(font)
                if isRepeat:
                        weight = 10.0 / faq
                        #4 / self.repeatList.index(self.parent.water_text_repeat)
                        for x in range(self.Base / 2, 2 * img.size[0], textwidth + int(45 * weight)):
                                for y in range(self.Base / 2, 2 * img.size[1], textheight + int(60 * weight)):
                                        img.watermark(imgTop, 1 - percentage, x, y)
                else:
                        img.watermark(imgTop, 1 - percentage, xy[0], xy[1])
                return self.parent.wand2pil(img, 'PNG')

        def isMousePos(self, textwidth, textheight, w, h):
                # check if mouse position is set manually?
                isPosSet = True
                if self.textRepeatCtrl.GetValue() == u'һ�� -> �����ָ��λ��':
                        if self.parent.startPoint:
                                left = self.parent.startPoint
                                scale = self.parent.scale
                                xy = (int(left.x * scale), int(left.y * scale))
                        else:
                                wx.MessageBox("����û��ָ��ˮӡλ����:-)��", "С����", wx.OK | wx.ICON_INFORMATION)
                                isPosSet = False
                                xy = None
                else: 
                        self.PositionDict = self._PositionMap(w, h, textwidth, textheight)                 
                        xy = self.PositionDict[self.fontPosition]
                return isPosSet, xy
        
        def GetWaterMarkText(self, filename):
                """
                get corresponding water mark
                """
                base = os.path.basename(filename) # only the filename
                if self.txtEnabled: return self.text
                elif self.m_radioPicName.GetValue(): return base
                elif self.m_radioFileList.GetValue():
                        # process text file or excel file # include self.text also
                        if self.excel_pic_names and self.listFile:
                                return self.text + '\n' + self.listFile[self.excel_pic_names.index(base.lower())]
                        else:
                            return self.text + '\n' + self.listFile[self.parent.picPaths.index(filename)]
                else: return None
                       
        #-----------------------------------------------------------
        def _im_mark(self, filename, istemp = True, im = None):
                """
                ���������ķ���, ����
                    filename : Դ�ļ���
                    im: ����Imageʵ��
                """
                
                # Check if we need to continue
                if (self.m_radioPic.GetValue() and self.m_textCtrlPath.GetValue() == "") or (self.m_radioFileList.GetValue() and self.m_fileListCtrlPath.GetValue() == "") or \
                (self.m_radioTxt.GetValue() and self.text == ""):
                        #print "invalid"
                        return
                
                ##############################
                # update parent parameters
                # 0 = no repeat; 1 = repeat
                self.parent.water_text_repeat = self.textRepeatCtrl.GetValue()
                self.parent.water_text_angle = self.textRotateCtrl.GetValue()
                self.parent.text_trans = self.transCtrl.GetValue()
                ANGLE = self.parent.water_text_angle
                isRepeat = not (self.parent.water_text_repeat.startswith( u'һ��'))
                faq = self.m_sliderFaq.GetValue()
                percentage = self.transCtrl.GetValue() * 0.01 
                ##############################
                
                # ����ˮӡ��Դ
                text = self.GetWaterMarkText(filename) #self.text
                #print filename, text
                if isinstance(text, str): text = unicode(text, 'cp936')
                
                picEnabled = self.m_radioPic.GetValue()
                if not picEnabled:
                
                        FONT = self.fontFile  
                        
                        # �趨������ɫ��͸���� -------------#
                        r,g,b = self.fontColour
                        
                        a = int( percentage * 255)#ȡ�ٷֱ�
                        self.finalColourWithAlpha = (r,g,b, a)
                        #-----------------------------------#                                                

                        cor_fill = self.finalColourWithAlpha
                        
                        if im:  img = im.convert("RGBA") # it is actual processing
                        
                        # Check if it is single mode
                        elif self.parent.singleMode and self.parent.im_preview:
                                img = self.parent.im_preview
                        else:   img = self.parent.imageOpen(filename).convert("RGBA") #
                        
                        # create a new water mark image
                        if isRepeat:                        
                                watermark = Image.new("RGBA", (2 * img.size[0], 2 * img.size[1]))
                        else:
                                watermark = Image.new("RGBA", (img.size[0], img.size[1]))

                        draw = ImageDraw.ImageDraw(watermark, "RGBA")
                        font = ImageFont.truetype(FONT, self.fontSize)
        
                        nexttextwidth, nexttextheight = font.getsize(text)
                        textwidth, textheight = nexttextwidth, nexttextheight
                        
                        #ȡ���꿪ʼ---------------------------------
                        w, h = watermark.size
                        
                        isPosSet, xy = self.isMousePos(textwidth, textheight, w, h)
                        if not isPosSet: return
        
                        #ȡ�������---------------------------------
                        
                        #draw.setfont(font)
                        if isRepeat:
                                #weight = 4 / self.repeatList.index(self.parent.water_text_repeat)
                                if faq == 1:
                                    r = math.radians(ANGLE)
                                    #
                                    xy = self.Base/2, self.Base/2
                                    for line in text.split('\n'):
                                        draw.text(xy, line, fill=cor_fill, font = font)
                                        xy = xy[0], xy[1] + textheight * 1.5
                                    text_list = text.split('\n')
                                    text_list.sort(key = lambda s: len(s))
                                    max_textwidth = font.getsize(text_list[-1])[0]
                                    rot_img = watermark.rotate(ANGLE, expand = 1)
                                    watermark = rot_img.crop((self.Base/2, int(self.Base /2 + math.sin(r) * (img.size[0] * 2 - max_textwidth - 15 )), (self.Base/2 + img.size[0]), int(self.Base /2 + math.sin(r) * ( img.size[0] * 2 - max_textwidth - 15) + img.size[1]) ) )
                                   
                                else:
                                    weight = 10.0 / faq
                                    for x in range(self.Base / 2, 2 * img.size[0], textwidth + int(45 * weight)):
                                            for y in range(self.Base / 2, 2 * img.size[1], textheight + int(60 * weight)):
                                                    draw.text((x,y), text, fill=cor_fill, font=font)
                                    rot_img = watermark.rotate(ANGLE, expand = 0)
                                    watermark = rot_img.crop((img.size[0]/2, img.size[1]/2, (img.size[0] * 3)/2, (img.size[1] * 3)/2) )                       
                        else:
                            for line in text.split('\n'):
                                draw.text(xy, line, fill=cor_fill, font = font)
                                xy = xy[0], xy[1] + textheight * 1.5
                                
                        # PIL merge of two images with alpha channels
                        img = Image.alpha_composite(img, watermark)
                else:
                        img = self.pic_watermark(isRepeat, ANGLE, percentage, filename, faq, im ) 
                if img:
                        if istemp:
                                ext = os.path.splitext(filename)[1]
                                self.parent.imgPreview(img, ext)                                            
                        else:
                                return img
                             
        #------------------------------------------------------------
        def _PositionMap(self,imgW, imgH, fontW, fontH):
                """
                ����ˮӡ���꺯��,����һ��PositionDict���ֵ�,������.����:
                   imgW : ͼƬ���
                   imgH : ͼƬ�߶�
                   fontW: ˮӡ���ֿ��
                   fontH: ˮӡ���ָ߶� 
                """
                #����ڶ��е�X����
                MX = (imgW-fontW)/2
                #��������е�X����
                RX = imgW - fontW - self.Base
                #����ڶ��е�Y���� 
                MY = (imgH-fontH)/2
                #����������²���Y����
                BY = imgH-fontH-self.Base
                PositionDict = {
                                 1:(self.Base, self.Base),#����
                                 2:(MX, self.Base),#����
                                 3:(RX, self.Base),#����
                                 4:(self.Base, MY),#����
                                 5:(MX, MY),#����
                                 6:(RX, MY),#����
                                 7:(self.Base, BY),#����
                                 8:(MX, BY),#����
                                 9:(RX, BY)#����
                                 }
                return PositionDict
        #-----------------------------------------------------------
        def OnRadioButton(self, event):
                # ���ˮӡλ��
                if self.LeftTop.GetValue():
                    self.fontPosition = 1#����
                elif self.CenterTop.GetValue():
                    self.fontPosition = 2#����
                elif self.RightTop.GetValue():
                    self.fontPosition = 3#����
                elif self.LeftCenter.GetValue():
                    self.fontPosition = 4#����
                elif self.Center.GetValue():
                    self.fontPosition = 5#����
                elif self.RightCenter.GetValue():
                    self.fontPosition = 6#����
                elif self.LeftBotm.GetValue():
                    self.fontPosition = 7#����
                elif self.CenterBotm.GetValue():
                    self.fontPosition = 8#����
                elif self.RightBotm.GetValue():
                    self.fontPosition = 9#����
                event.Skip()
        #------------------------------------------------------------
        def OnCancel(self, event):                
                #self.parent.isCancel = True
                #self.parent.picGUI.m_buttonWaterMark.SetValue(False)
                self.Hide() #��Ҫ����ɾ��
        #-------------------------------------------------------------
        def OnOK(self, event):   # preview             
                img = self.parent.getCurJpg()
                self.parent.current_dpi = self.parent.getSettingDPI(img)
                self._im_mark(img)
                #self.Show(False)
        #-----------------------------------------------------------
        def OnReset(self, event):
                self.parent.imgReload()
                
class ShowListDlg ( wx.Dialog ):
        
        def __init__( self, parent ):
                wx.Dialog.__init__ ( self, parent, id = wx.ID_ANY, title = u"��ǰ�ļ��б�", pos = wx.DefaultPosition, size = wx.Size( 587,530 ), style = wx.DEFAULT_DIALOG_STYLE|wx.MAXIMIZE_BOX|wx.MINIMIZE_BOX | wx.RESIZE_BORDER )
                
                sizer = wx.BoxSizer(wx.VERTICAL)
                panel = wx.Panel(self, wx.ID_ANY)
                self.index = 0
                
                self.m_listCtrlImgList = wx.ListCtrl( panel, wx.ID_ANY, wx.DefaultPosition, wx.Size(-1, 400), wx.BORDER_SUNKEN|wx.LC_REPORT )
                self.m_listCtrlImgList.InsertColumn(0, u'���')                
                self.m_listCtrlImgList.InsertColumn(1, u'����', width = 100)
                self.m_listCtrlImgList.InsertColumn(2, u'·��', width = 700)
                
                btn = wx.Button(panel, label=u"�ر�")
                btn.Bind(wx.EVT_BUTTON, self.close)
                
                sizer.Add(self.m_listCtrlImgList, 0, wx.ALL|wx.EXPAND, 5)
                sizer.Add(btn, 0, wx.ALL|wx.CENTER, 5)
                panel.SetSizer(sizer)

        # Virtual event handlers, overide them in your derived class
        def close ( self, event ):
                self.Hide()
                event.Skip()
                
        def addlines ( self, fileList ):
                while self.index < len(fileList):
                        path = os.path.dirname(fileList[self.index])
                        fname = os.path.basename(fileList[self.index])
                        self.m_listCtrlImgList.InsertStringItem(self.index, '%i' % (self.index + 1))
                        self.m_listCtrlImgList.SetStringItem(self.index, 1, fname)
                        self.m_listCtrlImgList.SetStringItem(self.index, 2, path)
                        self.index += 1